import requests
import concurrent.futures
import argparse
import csv
import sys
import time
import urllib3
from urllib.parse import urlparse
import socket
import re
from datetime import datetime
from collections import Counter
import os

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

class URLValidator:
    """
    Implementation of systematic subdomain operational status analysis protocol
    utilizing parallel connection establishment methodologies.
    """
    
    def __init__(self, timeout=8, workers=25, verify_ssl=False, follow_redirects=True, 
                 check_both_protocols=True):
        """
        Initialize the validator with operational parameters.
        
        Parameters:
            timeout (int): Connection establishment timeout threshold in seconds
            workers (int): Maximum concurrent execution threads
            verify_ssl (bool): SSL certificate validation parameter
            follow_redirects (bool): HTTP redirect traversal parameter
            check_both_protocols (bool): Check both HTTP and HTTPS for each URL
        """
        self.timeout = timeout
        self.max_workers = workers
        self.verify_ssl = verify_ssl
        self.follow_redirects = follow_redirects
        self.check_both_protocols = check_both_protocols
        self.session = self._initialize_session()
        
        # Disable SSL warnings if verification is disabled
        if not verify_ssl:
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    
    def _initialize_session(self):
        """
        Establish optimized HTTP session with standardized request parameters.
        
        Returns:
            requests.Session: Configured session object
        """
        session = requests.Session()
        session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.5',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
        })
        return session
    
    def validate_urls_from_file(self, file_path):
        """
        Process URL collection from file input source.
        
        Parameters:
            file_path (str): Path to file containing URL entries
            
        Returns:
            list: Collection of URL validation result objects
        """
        try:
            with open(file_path, 'r') as file:
                urls = [line.strip() for line in file if line.strip()]
            
            print(f"[+] Loaded {len(urls)} URLs from file")
            return self.validate_urls(urls)
        except FileNotFoundError:
            print(f"[-] Error: Input file not found - {file_path}")
            sys.exit(1)
    
    def validate_urls(self, urls):
        """
        Execute parallel URL validation protocol against specified endpoints.
        
        Parameters:
            urls (list): Collection of URL endpoint specifications
            
        Returns:
            list: Collection of URL validation result objects
        """
        # Normalize URL specifications
        normalized_urls = []
        for url in urls:
            # Ensure protocol specification
            if not url.startswith(('http://', 'https://')):
                if self.check_both_protocols:
                    normalized_urls.append(f"https://{url}")
                    normalized_urls.append(f"http://{url}")
                else:
                    # Only check HTTPS by default if only checking one protocol
                    normalized_urls.append(f"https://{url}")
            else:
                normalized_urls.append(url)
        
        # Remove duplicates while preserving order
        unique_urls = []
        seen = set()
        for url in normalized_urls:
            if url not in seen:
                unique_urls.append(url)
                seen.add(url)
        
        # Display operational parameters
        print(f"[+] Validating {len(unique_urls)} URLs with {self.max_workers} concurrent workers")
        print(f"[+] Timeout: {self.timeout} seconds | Follow redirects: {self.follow_redirects} | Verify SSL: {self.verify_ssl}")
        
        results = []
        start_time = time.time()
        
        # Implement parallel execution framework
        with concurrent.futures.ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            future_to_url = {executor.submit(self.check_url, url): url for url in unique_urls}
            
            # Process completed validation tasks
            for i, future in enumerate(concurrent.futures.as_completed(future_to_url)):
                url = future_to_url[future]
                try:
                    result = future.result()
                    results.append(result)
                    
                    # Display progress indicator
                    sys.stdout.write(f"\r[+] Progress: {i+1}/{len(unique_urls)} URLs processed")
                    sys.stdout.flush()
                except Exception as e:
                    print(f"\n[-] Error processing {url}: {str(e)}")
        
        elapsed_time = time.time() - start_time
        print(f"\n[+] Validation completed in {elapsed_time:.2f} seconds")
        
        return results
    
    def check_url(self, url):
        """
        Execute comprehensive validation protocol against specified endpoint.
        
        Parameters:
            url (str): URL endpoint specification
            
        Returns:
            dict: Structured validation result object
        """
        result = {
            'url': url,
            'status': 'inactive',
            'response_code': None,
            'response_time': None,
            'content_length': None,
            'title': None,
            'server': None,
            'ip_address': None,
            'redirect_url': None,
            'error': None,
            'protocol': urlparse(url).scheme
        }
        
        try:
            # Extract domain from URL
            parsed_url = urlparse(url)
            domain = parsed_url.netloc
            result['domain'] = domain
            
            # Resolve DNS first for faster error detection
            try:
                ip_address = socket.gethostbyname(domain)
                result['ip_address'] = ip_address
            except socket.gaierror:
                result['error'] = 'DNS resolution failed'
                return result
            
            # Execute HTTP request with timeout parameter
            start_time = time.time()
            response = self.session.get(
                url,
                timeout=self.timeout,
                verify=self.verify_ssl,
                allow_redirects=self.follow_redirects
            )
            response_time = time.time() - start_time
            
            # Process response metadata
            result['status'] = 'active'
            result['response_code'] = response.status_code
            result['response_time'] = round(response_time * 1000)  # Convert to milliseconds
            result['content_length'] = len(response.content)
            result['server'] = response.headers.get('Server', 'Unknown')
            
            # Extract title from HTML content
            if 'text/html' in response.headers.get('Content-Type', ''):
                title_match = re.search(r'<title>(.*?)</title>', response.text, re.IGNORECASE | re.DOTALL)
                if title_match:
                    result['title'] = title_match.group(1).strip()
            
            # Record redirect destination
            if self.follow_redirects and response.history:
                result['redirect_url'] = response.url
        
        except requests.exceptions.SSLError:
            result['error'] = 'SSL error'
        
        except requests.exceptions.ConnectionError:
            result['error'] = 'Connection error'
        
        except requests.exceptions.Timeout:
            result['error'] = 'Timeout'
        
        except requests.exceptions.TooManyRedirects:
            result['error'] = 'Too many redirects'
        
        except requests.exceptions.RequestException as e:
            result['error'] = f'Request error: {str(e)}'
        
        return result
    
    def analyze_results(self, results):
        """
        Perform statistical analysis of validation result collection.
        
        Parameters:
            results (list): Collection of validation result objects
            
        Returns:
            dict: Structured statistical analysis object
        """
        # Filter results by status
        active_urls = [r for r in results if r['status'] == 'active']
        inactive_urls = [r for r in results if r['status'] == 'inactive']
        
        # Analyze response code distribution
        response_codes = Counter([r['response_code'] for r in active_urls if r['response_code']])
        
        # Analyze error distribution
        error_types = Counter([r['error'] for r in inactive_urls if r['error']])
        
        # Calculate response time statistics
        response_times = [r['response_time'] for r in active_urls if r['response_time']]
        avg_response_time = sum(response_times) / len(response_times) if response_times else 0
        
        # Calculate content length statistics
        content_lengths = [r['content_length'] for r in active_urls if r['content_length']]
        avg_content_length = sum(content_lengths) / len(content_lengths) if content_lengths else 0
        
        # Analyze protocol distribution
        protocols = Counter([r['protocol'] for r in results if r['protocol']])
        
        # Analyze domain distribution (get unique domains)
        domains = set([r['domain'] for r in results if 'domain' in r])
        
        return {
            'total_urls': len(results),
            'active_count': len(active_urls),
            'inactive_count': len(inactive_urls),
            'response_codes': response_codes,
            'error_types': error_types,
            'avg_response_time': avg_response_time,
            'avg_content_length': avg_content_length,
            'protocols': protocols,
            'unique_domains': len(domains)
        }
    
    def export_results_to_csv(self, results, output_file):
        """
        Export validation results to structured CSV format.
        
        Parameters:
            results (list): Collection of validation result objects
            output_file (str): Output file path specification
        """
        fieldnames = [
            'url', 'domain', 'protocol', 'status', 'response_code', 'response_time', 
            'content_length', 'title', 'server', 'ip_address', 
            'redirect_url', 'error'
        ]
        
        with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(results)
            
        print(f"[+] Results exported to CSV: {output_file}")
    
    def export_results_to_excel(self, results, output_file):
        """
        Export validation results to formatted Excel workbook.
        
        Parameters:
            results (list): Collection of validation result objects
            output_file (str): Output file path specification
        """
        if not EXCEL_AVAILABLE:
            print("[-] Excel export requires openpyxl library. Install with: pip install openpyxl")
            print(f"[+] Falling back to CSV export: {output_file.replace('.xlsx', '.csv')}")
            self.export_results_to_csv(results, output_file.replace('.xlsx', '.csv'))
            return
            
        # Create workbook and sheets
        wb = openpyxl.Workbook()
        
        # Create main results sheet
        ws_results = wb.active
        ws_results.title = "URL Validation Results"
        
        # Create summary sheet
        ws_summary = wb.create_sheet("Summary")
        
        # Define header fields
        headers = [
            'URL', 'Domain', 'Protocol', 'Status', 'Response Code', 
            'Response Time (ms)', 'Content Length', 'Page Title', 
            'Server', 'IP Address', 'Redirect URL', 'Error'
        ]
        
        # Define column widths
        column_widths = [50, 30, 10, 10, 15, 20, 15, 40, 20, 15, 50, 30]
        
        # Add headers to results sheet
        for col_num, header in enumerate(headers, 1):
            cell = ws_results.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
            # Set column width
            ws_results.column_dimensions[get_column_letter(col_num)].width = column_widths[col_num-1]
        
        # Add data to results sheet
        for row_num, result in enumerate(results, 2):
            # URL
            ws_results.cell(row=row_num, column=1).value = result['url']
            
            # Domain
            ws_results.cell(row=row_num, column=2).value = result.get('domain', '')
            
            # Protocol
            ws_results.cell(row=row_num, column=3).value = result.get('protocol', '')
            
            # Status
            status_cell = ws_results.cell(row=row_num, column=4)
            status_cell.value = result['status']
            if result['status'] == 'active':
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            
            # Response Code
            code_cell = ws_results.cell(row=row_num, column=5)
            code_cell.value = result['response_code']
            if result['response_code'] in [200, 301, 302]:
                code_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif result['response_code'] and 400 <= result['response_code'] < 500:
                code_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif result['response_code'] and result['response_code'] >= 500:
                code_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            
            # Response Time
            ws_results.cell(row=row_num, column=6).value = result['response_time']
            
            # Content Length
            ws_results.cell(row=row_num, column=7).value = result['content_length']
            
            # Title
            ws_results.cell(row=row_num, column=8).value = result.get('title', '')
            
            # Server
            ws_results.cell(row=row_num, column=9).value = result.get('server', '')
            
            # IP Address
            ws_results.cell(row=row_num, column=10).value = result.get('ip_address', '')
            
            # Redirect URL
            ws_results.cell(row=row_num, column=11).value = result.get('redirect_url', '')
            
            # Error
            error_cell = ws_results.cell(row=row_num, column=12)
            error_cell.value = result.get('error', '')
            if result.get('error'):
                error_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        
        # Create Autofilter
        ws_results.auto_filter.ref = ws_results.dimensions
        
        # Freeze header row
        ws_results.freeze_panes = 'A2'
        
        # Prepare summary data
        # Count active/inactive URLs
        active_count = len([r for r in results if r['status'] == 'active'])
        inactive_count = len([r for r in results if r['status'] == 'inactive'])
        
        # Count response codes
        response_codes = Counter([r['response_code'] for r in results if r['response_code']])
        
        # Count error types
        error_types = Counter([r['error'] for r in results if r['error']])
        
        # Add summary data
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 15
        
        # Title
        cell = ws_summary.cell(row=1, column=1)
        cell.value = "URL Validation Summary"
        cell.font = Font(bold=True, size=14)
        ws_summary.merge_cells('A1:C1')
        cell.alignment = Alignment(horizontal='center')
        
        # Overall Statistics
        ws_summary.cell(row=3, column=1).value = "Total URLs Validated:"
        ws_summary.cell(row=3, column=2).value = len(results)
        ws_summary.cell(row=3, column=1).font = Font(bold=True)
        
        ws_summary.cell(row=4, column=1).value = "Active URLs:"
        ws_summary.cell(row=4, column=2).value = active_count
        ws_summary.cell(row=4, column=3).value = f"{active_count/len(results)*100:.1f}%"
        ws_summary.cell(row=4, column=1).font = Font(bold=True)
        
        ws_summary.cell(row=5, column=1).value = "Inactive URLs:"
        ws_summary.cell(row=5, column=2).value = inactive_count
        ws_summary.cell(row=5, column=3).value = f"{inactive_count/len(results)*100:.1f}%"
        ws_summary.cell(row=5, column=1).font = Font(bold=True)
        
        # Response Code Distribution
        ws_summary.cell(row=7, column=1).value = "Response Code Distribution"
        ws_summary.cell(row=7, column=1).font = Font(bold=True)
        
        row = 8
        for code, count in sorted(response_codes.items()):
            ws_summary.cell(row=row, column=1).value = f"Status {code}:"
            ws_summary.cell(row=row, column=2).value = count
            if active_count > 0:
                ws_summary.cell(row=row, column=3).value = f"{count/active_count*100:.1f}%"
            row += 1
        
        # Error Type Distribution
        ws_summary.cell(row=row+1, column=1).value = "Error Type Distribution"
        ws_summary.cell(row=row+1, column=1).font = Font(bold=True)
        
        row += 2
        for error, count in sorted(error_types.items()):
            ws_summary.cell(row=row, column=1).value = error
            ws_summary.cell(row=row, column=2).value = count
            if inactive_count > 0:
                ws_summary.cell(row=row, column=3).value = f"{count/inactive_count*100:.1f}%"
            row += 1
        
        # Add timestamp
        ws_summary.cell(row=row+2, column=1).value = "Report Generated:"
        ws_summary.cell(row=row+2, column=2).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_summary.cell(row=row+2, column=1).font = Font(bold=True)
        
        # Save workbook
        try:
            wb.save(output_file)
            print(f"[+] Results exported to Excel: {output_file}")
        except PermissionError:
            print(f"[-] Unable to write to {output_file} - file may be in use")
            alternate_file = f"{os.path.splitext(output_file)[0]}_new.xlsx"
            wb.save(alternate_file)
            print(f"[+] Results exported to alternate file: {alternate_file}")
    
    def display_results_summary(self, results, analysis):
        """
        Generate formatted console output of validation results.
        
        Parameters:
            results (list): Collection of validation result objects
            analysis (dict): Statistical analysis object
        """
        print("\n" + "=" * 70)
        print(f"URL VALIDATION SUMMARY")
        print("=" * 70)
        print(f"Total URLs processed: {analysis['total_urls']}")
        print(f"Active URLs: {analysis['active_count']} ({analysis['active_count']/analysis['total_urls']*100:.1f}%)")
        print(f"Inactive URLs: {analysis['inactive_count']} ({analysis['inactive_count']/analysis['total_urls']*100:.1f}%)")
        
        # Display unique domains count if check_both_protocols is enabled
        if self.check_both_protocols:
            print(f"Unique domains: {analysis['unique_domains']}")
        
        # Display protocol distribution if check_both_protocols is enabled
        if self.check_both_protocols and analysis['protocols']:
            print("\nProtocol Distribution:")
            for protocol, count in sorted(analysis['protocols'].items()):
                print(f"  {protocol.upper()}: {count} URLs ({count/analysis['total_urls']*100:.1f}%)")
        
        # Display response code distribution
        if analysis['response_codes']:
            print("\nResponse Code Distribution:")
            for code, count in sorted(analysis['response_codes'].items()):
                print(f"  {code}: {count} URLs ({count/analysis['active_count']*100:.1f}%)")
        
        # Display error type distribution
        if analysis['error_types']:
            print("\nError Type Distribution:")
            for error, count in sorted(analysis['error_types'].items()):
                print(f"  {error}: {count} URLs ({count/analysis['inactive_count']*100:.1f}%)")
        
        # Display performance metrics
        if analysis['avg_response_time'] > 0:
            print(f"\nAverage Response Time: {analysis['avg_response_time']:.1f} ms")
        
        print("\n" + "=" * 70)
        print("ACTIVE URLs (Sample)")
        print("=" * 70)
        
        # Display active URLs sample
        active_urls = [r for r in results if r['status'] == 'active']
        for r in active_urls[:10]:  # Show first 10 active URLs
            print(f"{r['url']} - Code: {r['response_code']} - Time: {r['response_time']} ms")
        
        if len(active_urls) > 10:
            print(f"... and {len(active_urls) - 10} additional active URLs")
        
        print("\n" + "=" * 70)
        print("INACTIVE URLs (Sample)")
        print("=" * 70)
        
        # Display inactive URLs sample
        inactive_urls = [r for r in results if r['status'] == 'inactive']
        for r in inactive_urls[:10]:  # Show first 10 inactive URLs
            print(f"{r['url']} - Error: {r['error']}")
        
        if len(inactive_urls) > 10:
            print(f"... and {len(inactive_urls) - 10} additional inactive URLs")


def main():
    """
    Primary execution function implementing command-line interface.
    """
    parser = argparse.ArgumentParser(
        description="URL Validation Protocol: Systematically assess subdomain operational status",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter
    )
    
    # Input source specification
    input_group = parser.add_mutually_exclusive_group(required=True)
    input_group.add_argument('-f', '--file', help='Input file containing URLs (one per line)')
    input_group.add_argument('-u', '--url', help='Single URL to validate')
    input_group.add_argument('-d', '--domain', help='Base domain for subdomain generation')
    
    # Output specification
    parser.add_argument('-o', '--output', help='Output file (.csv or .xlsx format)')
    
    # Execution parameters
    parser.add_argument('-t', '--timeout', type=int, default=8, help='Request timeout in seconds')
    parser.add_argument('-w', '--workers', type=int, default=25, help='Number of concurrent workers')
    parser.add_argument('-s', '--ssl', action='store_true', help='Verify SSL certificates')
    parser.add_argument('-r', '--no-redirect', action='store_true', help='Do not follow redirects')
    parser.add_argument('-p', '--single-protocol', action='store_true', 
                      help='Check only one protocol (HTTPS) instead of both HTTP and HTTPS')
    
    # Parse arguments
    args = parser.parse_args()
    
    # Initialize validator
    validator = URLValidator(
        timeout=args.timeout,
        workers=args.workers,
        verify_ssl=args.ssl,
        follow_redirects=not args.no_redirect,
        check_both_protocols=not args.single_protocol
    )
    
    # Process URLs based on input type
    if args.file:
        print(f"[+] Reading URLs from file: {args.file}")
        results = validator.validate_urls_from_file(args.file)
    elif args.url:
        print(f"[+] Validating single URL: {args.url}")
        results = validator.validate_urls([args.url])
    elif args.domain:
        print(f"[+] Generating common subdomains for: {args.domain}")
        # Common subdomain prefixes
        prefixes = [
            'www', 'app', 'api', 'blog', 'mail', 'remote', 'webmail', 'portal', 
            'admin', 'dev', 'staging', 'test', 'beta', 'dashboard', 'secure', 
            'shop', 'store', 'm', 'mobile', 'support', 'help', 'docs', 'login', 
            'auth', 'account', 'cdn', 'media', 'static', 'assets', 'news', 
            'forum', 'community', 'events', 'services', 'cloud', 'pay', 'status',
            'money', 'finance', 'bank', 'health', 'social', 'business', 'corp'
        ]
        urls = [f"{prefix}.{args.domain}" for prefix in prefixes]
        results = validator.validate_urls(urls)
    
    # Analyze results
    analysis = validator.analyze_results(results)
    
    # Display results summary
    validator.display_results_summary(results, analysis)
    
    # Export results based on file extension
    if args.output:
        if args.output.lower().endswith('.xlsx'):
            validator.export_results_to_excel(results, args.output)
        else:
            validator.export_results_to_csv(results, args.output)


if __name__ == "__main__":
    # Display execution banner
    print("\n" + "=" * 70)
    print("  SUBDOMAIN OPERATIONAL STATUS ANALYSIS PROTOCOL")
    print("=" * 70)
    
    start_time = time.time()
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[+] Execution initiated: {timestamp}\n")
    
    try:
        main()
    except KeyboardInterrupt:
        print("\n[-] Execution terminated by user")
    except Exception as e:
        print(f"\n[-] Execution error: {str(e)}")
    
    elapsed_time = time.time() - start_time
    print(f"\n[+] Total execution time: {elapsed_time:.2f} seconds")