#!/usr/bin/env python3
"""
DAST Scanner - Dynamic Application Security Testing Tool
This tool performs in-depth security scanning on web applications, identifies
vulnerable components, and generates professional Excel reports.
"""

import argparse
import logging
import os
import re
import sys
import time
from datetime import datetime
from urllib.parse import urlparse

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from zapv2 import ZAPv2

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler('dast_scanner.log')
    ]
)
logger = logging.getLogger('DAST-Scanner')

class DASTScanner:
    """Main class for DAST scanning operations"""
    
    def __init__(self, args):
        """Initialize the scanner with command line arguments"""
        self.args = args
        self.urls = []
        self.results = []
        self.zap = None
        self.output_file = args.output or f"security_scan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Colors for Excel report
        self.colors = {
            'header': 'CCE5FF',  # Light blue
            'high': 'FF9999',    # Light red
            'medium': 'FFCC99',  # Light orange
            'low': 'FFFFCC',     # Light yellow
            'info': 'E6F2FF',     # Very light blue
            'critical': 'FF6666' # Darker red for critical findings
        }
        
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36'
        })
        
        if args.zap_api_key and args.zap_proxy:
            self.init_zap()
    
    def init_zap(self):
        """Initialize ZAP API connection"""
        try:
            logger.info(f"Connecting to ZAP API at {self.args.zap_proxy}")
            self.zap = ZAPv2(apikey=self.args.zap_api_key, proxies={'http': self.args.zap_proxy, 'https': self.args.zap_proxy})
            logger.info(f"Successfully connected to ZAP API: {self.zap.core.version}")
        except Exception as e:
            logger.error(f"Failed to connect to ZAP API: {e}")
            if self.args.require_zap:
                sys.exit(1)
    
    def read_urls(self):
        """Read URLs from input file"""
        if not os.path.exists(self.args.input_file):
            logger.error(f"Input file not found: {self.args.input_file}")
            sys.exit(1)
            
        with open(self.args.input_file, 'r') as f:
            for line in f:
                url = line.strip()
                if url and not url.startswith('#'):  # Skip empty lines and comments
                    # Add http:// prefix if not present
                    if not url.startswith(('http://', 'https://')):
                        url = 'https://' + url
                    self.urls.append(url)
        
        logger.info(f"Loaded {len(self.urls)} URLs from {self.args.input_file}")
        if not self.urls:
            logger.error("No valid URLs found in input file")
            sys.exit(1)
    
    def scan_with_zap(self, url):
        """Perform scanning using OWASP ZAP"""
        if not self.zap:
            return []
            
        try:
            logger.info(f"Starting ZAP active scan on {url}")
            
            # Access the target
            logger.info("Accessing target...")
            self.zap.urlopen(url)
            self.zap.spider.scan(url)
            
            # Wait for spider to complete
            logger.info("Spider running...")
            while int(self.zap.spider.status) < 100:
                logger.info(f"Spider progress: {self.zap.spider.status}%")
                time.sleep(2)
            
            logger.info("Spider complete")
            
            # Run the active scan
            scan_id = self.zap.ascan.scan(url)
            logger.info(f"Active scan started with ID {scan_id}")
            
            # Wait for active scan to complete
            while int(self.zap.ascan.status(scan_id)) < 100:
                logger.info(f"Active scan progress: {self.zap.ascan.status(scan_id)}%")
                time.sleep(5)
            
            logger.info("Active scan complete")
            
            # Get scan results
            alerts = self.zap.core.alerts(url)
            logger.info(f"Found {len(alerts)} vulnerabilities")
            
            findings = []
            for alert in alerts:
                findings.append({
                    'URL': url,
                    'Type': 'ZAP Scan',
                    'Name': alert.get('name', 'Unknown'),
                    'Description': alert.get('description', ''),
                    'Risk': alert.get('risk', 'Info'),
                    'Confidence': alert.get('confidence', 'Low'),
                    'Path': alert.get('url', ''),
                    'Solution': alert.get('solution', ''),
                    'References': alert.get('reference', '')
                })
            
            return findings
            
        except Exception as e:
            logger.error(f"Error during ZAP scan: {e}")
            return []
    
    def analyze_components(self, url):
        """Analyze web components for known vulnerabilities"""
        findings = []
        
        try:
            logger.info(f"Analyzing components for {url}")
            response = self.session.get(url, timeout=30, verify=not self.args.ignore_ssl)
            if response.status_code != 200:
                logger.warning(f"Failed to access {url}, status: {response.status_code}")
                return findings
            
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Extract JavaScript libraries
            js_libs = []
            for script in soup.find_all('script', src=True):
                src = script['src']
                js_libs.append(src)
                
                # Try to identify library and version
                lib_name, version = self.identify_library(src)
                if lib_name:
                    risk_level = self.check_component_vulnerability(lib_name, version)
                    if risk_level:
                        findings.append({
                            'URL': url,
                            'Type': 'Component Analysis',
                            'Name': f"Vulnerable JavaScript Library: {lib_name}",
                            'Description': f"Detected {lib_name} version {version} which has known vulnerabilities",
                            'Risk': risk_level,
                            'Confidence': 'Medium',
                            'Path': src,
                            'Solution': f"Update {lib_name} to the latest version",
                            'References': 'https://snyk.io/vuln'
                        })
            
            # Extract CSS libraries
            css_libs = []
            for link in soup.find_all('link', rel='stylesheet', href=True):
                href = link['href']
                css_libs.append(href)
                
                # Check for vulnerable CSS framework
                lib_name, version = self.identify_library(href)
                if lib_name:
                    risk_level = self.check_component_vulnerability(lib_name, version)
                    if risk_level:
                        findings.append({
                            'URL': url,
                            'Type': 'Component Analysis',
                            'Name': f"Vulnerable CSS Framework: {lib_name}",
                            'Description': f"Detected {lib_name} version {version} which has known vulnerabilities",
                            'Risk': risk_level,
                            'Confidence': 'Medium',
                            'Path': href,
                            'Solution': f"Update {lib_name} to the latest version",
                            'References': 'https://snyk.io/vuln'
                        })
            
            # Check server headers for information disclosure
            server = response.headers.get('Server')
            if server:
                findings.append({
                    'URL': url,
                    'Type': 'Information Disclosure',
                    'Name': 'Server Information Disclosure',
                    'Description': f"Server header reveals: {server}",
                    'Risk': 'Low',
                    'Confidence': 'High',
                    'Path': url,
                    'Solution': 'Configure server to hide version information',
                    'References': 'https://owasp.org/www-project-web-security-testing-guide/latest/4-Web_Application_Security_Testing/02-Configuration_and_Deployment_Management_Testing/03-Test_File_Extensions_Handling_for_Sensitive_Information'
                })
            
            # Extract meta data
            for meta in soup.find_all('meta'):
                name = meta.get('name', '').lower()
                if name in ['generator', 'application-name']:
                    content = meta.get('content', '')
                    if content:
                        findings.append({
                            'URL': url,
                            'Type': 'Information Disclosure',
                            'Name': 'Technology Disclosure',
                            'Description': f"Meta tag reveals: {name}={content}",
                            'Risk': 'Info',
                            'Confidence': 'Medium',
                            'Path': url,
                            'Solution': 'Consider removing version information from meta tags',
                            'References': ''
                        })
            
            # Check content-security-policy
            csp = response.headers.get('Content-Security-Policy')
            if not csp:
                findings.append({
                    'URL': url,
                    'Type': 'Security Headers',
                    'Name': 'Missing Content-Security-Policy',
                    'Description': "Content Security Policy header is not set",
                    'Risk': 'Medium',
                    'Confidence': 'High',
                    'Path': url,
                    'Solution': 'Implement a Content Security Policy',
                    'References': 'https://developer.mozilla.org/en-US/docs/Web/HTTP/CSP'
                })
            
            # Check X-Frame-Options
            x_frame = response.headers.get('X-Frame-Options')
            if not x_frame:
                findings.append({
                    'URL': url,
                    'Type': 'Security Headers',
                    'Name': 'Missing X-Frame-Options',
                    'Description': "X-Frame-Options header is not set, potential clickjacking risk",
                    'Risk': 'Medium',
                    'Confidence': 'High',
                    'Path': url,
                    'Solution': 'Implement X-Frame-Options header',
                    'References': 'https://developer.mozilla.org/en-US/docs/Web/HTTP/Headers/X-Frame-Options'
                })
                
            # Also check for outdated CMS/frameworks based on meta tags
            cms_patterns = {
                'WordPress': r'<meta\s+name="generator"\s+content="WordPress\s+(\d+\.\d+(\.\d+)?).*?"',
                'Drupal': r'<meta\s+name="Generator"\s+content="Drupal\s+(\d+\.\d+(\.\d+)?).*?"',
                'Joomla': r'<meta\s+name="generator"\s+content="Joomla!\s+(\d+\.\d+(\.\d+)?).*?"',
                'Magento': r'<meta\s+name="generator"\s+content="Magento\s+(\d+\.\d+(\.\d+)?).*?"',
                'Ghost': r'<meta\s+name="generator"\s+content="Ghost\s+(\d+\.\d+(\.\d+)?).*?"',
                'Shopify': r'<meta\s+name="generator"\s+content="Shopify\s+(\d+\.\d+(\.\d+)?).*?"'
            }
            
            for cms, pattern in cms_patterns.items():
                match = re.search(pattern, response.text)
                if match:
                    version = match.group(1)
                    findings.append({
                        'URL': url,
                        'Type': 'CMS Detection',
                        'Name': f"Detected {cms}",
                        'Description': f"Detected {cms} version {version} which may have known vulnerabilities",
                        'Risk': 'Info',
                        'Confidence': 'Medium',
                        'Path': url,
                        'Solution': f"Keep {cms} updated to the latest version and remove version disclosure",
                        'References': f"https://www.cvedetails.com/vendor/search.php?vendor={cms.lower()}"
                    })
            
            logger.info(f"Component analysis complete for {url}, found {len(findings)} issues")
            return findings
            
        except Exception as e:
            logger.error(f"Error during component analysis: {e}")
            return findings
    
    def identify_library(self, path):
        """Attempt to identify library name and version from path"""
        # Sample patterns to identify common libraries - expanded list
        patterns = [
            (r'jquery[.-](\d+\.\d+\.\d+)', 'jQuery'),
            (r'bootstrap[.-](\d+\.\d+\.\d+)', 'Bootstrap'),
            (r'react[.-](\d+\.\d+\.\d+)', 'React'),
            (r'angular[.-](\d+\.\d+\.\d+)', 'Angular'),
            (r'vue[.-](\d+\.\d+\.\d+)', 'Vue.js'),
            (r'lodash[.-](\d+\.\d+\.\d+)', 'Lodash'),
            (r'moment[.-](\d+\.\d+\.\d+)', 'Moment.js'),
            (r'axios[.-](\d+\.\d+\.\d+)', 'Axios'),
            (r'ember[.-](\d+\.\d+\.\d+)', 'Ember.js'),
            (r'backbone[.-](\d+\.\d+\.\d+)', 'Backbone.js'),
            (r'foundation[.-](\d+\.\d+\.\d+)', 'Foundation'),
            (r'materialize[.-](\d+\.\d+\.\d+)', 'Materialize CSS'),
            (r'bulma[.-](\d+\.\d+\.\d+)', 'Bulma'),
            (r'tailwind[.-](\d+\.\d+\.\d+)', 'Tailwind CSS'),
            (r'three[.-](\d+\.\d+\.\d+)', 'Three.js'),
            (r'd3[.-](\d+\.\d+\.\d+)', 'D3.js'),
            (r'leaflet[.-](\d+\.\d+\.\d+)', 'Leaflet'),
            (r'chart[.-](\d+\.\d+\.\d+)', 'Chart.js'),
            (r'highcharts[.-](\d+\.\d+\.\d+)', 'Highcharts'),
            (r'tensorflow[.-](\d+\.\d+\.\d+)', 'TensorFlow.js'),
            (r'gsap[.-](\d+\.\d+\.\d+)', 'GSAP'),
            (r'sweetalert[.-](\d+\.\d+\.\d+)', 'SweetAlert')
        ]
        
        for pattern, lib_name in patterns:
            match = re.search(pattern, path, re.IGNORECASE)
            if match:
                return lib_name, match.group(1)
        
        return None, None
    
    def check_component_vulnerability(self, library, version):
        """Check if component has known vulnerabilities"""
        # Expanded list of vulnerable library versions (for demonstration)
        vulnerable_libs = {
            'jQuery': [
                ('1.0.0', '1.12.4', 'Medium'),
                ('2.0.0', '2.2.4', 'High'),
                ('3.0.0', '3.4.0', 'Medium')
            ],
            'Bootstrap': [
                ('2.0.0', '2.3.2', 'Medium'),
                ('3.0.0', '3.3.0', 'Low'),
                ('4.0.0', '4.3.1', 'Medium')
            ],
            'Angular': [
                ('1.0.0', '1.5.0', 'High'),
                ('1.5.9', '1.5.11', 'Medium')
            ],
            'React': [
                ('16.0.0', '16.2.0', 'Medium'),
                ('0.5.0', '0.14.0', 'High')
            ],
            'Lodash': [
                ('4.0.0', '4.17.15', 'High')
            ],
            'Vue.js': [
                ('2.0.0', '2.5.16', 'Medium')
            ],
            'Moment.js': [
                ('2.0.0', '2.19.2', 'Medium')
            ],
            'Axios': [
                ('0.15.0', '0.18.0', 'High')
            ]
        }
        
        if library in vulnerable_libs:
            for v_range in vulnerable_libs[library]:
                min_ver, max_ver, risk = v_range
                if self._version_in_range(version, min_ver, max_ver):
                    return risk
        
        return None
    
    def _version_in_range(self, version, min_version, max_version):
        """Check if version is in range (inclusive)"""
        try:
            version_parts = list(map(int, version.split('.')))
            min_parts = list(map(int, min_version.split('.')))
            max_parts = list(map(int, max_version.split('.')))
            
            # Pad with zeros if necessary
            max_length = max(len(version_parts), len(min_parts), len(max_parts))
            version_parts += [0] * (max_length - len(version_parts))
            min_parts += [0] * (max_length - len(min_parts))
            max_parts += [0] * (max_length - len(max_parts))
            
            # Check if version is in range
            return min_parts <= version_parts <= max_parts
        except:
            return False
    
    def scan_url(self, url):
        """Perform all scanning operations on a single URL"""
        logger.info(f"Starting scan for {url}")
        
        findings = []
        
        # Basic checks
        findings.extend(self.basic_security_checks(url))
        
        # Component analysis
        findings.extend(self.analyze_components(url))
        
        # ZAP scanning if enabled
        if self.zap:
            findings.extend(self.scan_with_zap(url))
        
        logger.info(f"Completed scan for {url}, found {len(findings)} issues")
        return findings
    
    def basic_security_checks(self, url):
        """Perform basic security checks without external tools"""
        findings = []
        try:
            # Check SSL/TLS
            parsed_url = urlparse(url)
            if parsed_url.scheme == 'https':
                try:
                    response = self.session.get(url, timeout=10, verify=True)
                except requests.exceptions.SSLError:
                    findings.append({
                        'URL': url,
                        'Type': 'SSL/TLS',
                        'Name': 'Invalid SSL Certificate',
                        'Description': 'The SSL certificate is invalid or self-signed',
                        'Risk': 'High',
                        'Confidence': 'High',
                        'Path': url,
                        'Solution': 'Install a valid SSL certificate from a trusted CA',
                        'References': 'https://www.digicert.com/ssl/'
                    })
            else:
                findings.append({
                    'URL': url,
                    'Type': 'SSL/TLS',
                    'Name': 'No SSL/TLS',
                    'Description': 'The site does not use SSL/TLS encryption',
                    'Risk': 'High',
                    'Confidence': 'High',
                    'Path': url,
                    'Solution': 'Implement HTTPS with a valid certificate',
                    'References': 'https://www.digicert.com/ssl/'
                })
                
            # Check for common sensitive files
            sensitive_paths = [
                '/robots.txt', 
                '/sitemap.xml', 
                '/.git/', 
                '/.env', 
                '/backup/', 
                '/wp-config.php',
                '/.htaccess',
                '/.svn/',
                '/.DS_Store',
                '/phpinfo.php',
                '/server-status',
                '/server-info',
                '/.bash_history',
                '/.ssh/',
                '/config.php',
                '/database.yml',
                '/credentials.txt',
                '/api-docs',
                '/swagger'
            ]
            base_url = f"{parsed_url.scheme}://{parsed_url.netloc}"
            
            for path in sensitive_paths:
                try:
                    test_url = base_url + path
                    resp = self.session.get(test_url, timeout=5, verify=not self.args.ignore_ssl)
                    if resp.status_code == 200:
                        findings.append({
                            'URL': url,
                            'Type': 'Information Disclosure',
                            'Name': f'Sensitive File Exposed: {path}',
                            'Description': f'The file {path} is accessible and may contain sensitive information',
                            'Risk': 'Medium',
                            'Confidence': 'Medium',
                            'Path': test_url,
                            'Solution': f'Restrict access to {path} or remove it if not needed',
                            'References': 'https://owasp.org/www-project-web-security-testing-guide/latest/4-Web_Application_Security_Testing/02-Configuration_and_Deployment_Management_Testing/03-Test_File_Extensions_Handling_for_Sensitive_Information'
                        })
                except:
                    pass
                    
            # Test for HTTP Methods
            try:
                options_resp = self.session.options(url, timeout=5, verify=not self.args.ignore_ssl)
                allowed_methods = options_resp.headers.get('Allow', '')
                if allowed_methods:
                    dangerous_methods = ['PUT', 'DELETE', 'TRACE']
                    for method in dangerous_methods:
                        if method in allowed_methods:
                            findings.append({
                                'URL': url,
                                'Type': 'Configuration',
                                'Name': f'Dangerous HTTP Method Allowed: {method}',
                                'Description': f'The server allows the {method} HTTP method which can be dangerous',
                                'Risk': 'Medium',
                                'Confidence': 'Medium',
                                'Path': url,
                                'Solution': f'Disable the {method} HTTP method if not required',
                                'References': 'https://owasp.org/www-project-web-security-testing-guide/latest/4-Web_Application_Security_Testing/02-Configuration_and_Deployment_Management_Testing/06-Test_HTTP_Methods'
                            })
            except:
                pass
                
            # Check security headers
            response = self.session.get(url, timeout=10, verify=not self.args.ignore_ssl)
            
            # Check Strict-Transport-Security (HSTS)
            hsts = response.headers.get('Strict-Transport-Security')
            if not hsts and parsed_url.scheme == 'https':
                findings.append({
                    'URL': url,
                    'Type': 'Security Headers',
                    'Name': 'Missing HSTS Header',
                    'Description': 'HTTP Strict Transport Security header is not set',
                    'Risk': 'Medium',
                    'Confidence': 'High',
                    'Path': url,
                    'Solution': 'Implement the Strict-Transport-Security header',
                    'References': 'https://developer.mozilla.org/en-US/docs/Web/HTTP/Headers/Strict-Transport-Security'
                })
            
            # Check X-XSS-Protection
            xss_protection = response.headers.get('X-XSS-Protection')
            if not xss_protection:
                findings.append({
                    'URL': url,
                    'Type': 'Security Headers',
                    'Name': 'Missing X-XSS-Protection',
                    'Description': 'X-XSS-Protection header is not set',
                    'Risk': 'Low',
                    'Confidence': 'High',
                    'Path': url,
                    'Solution': 'Implement the X-XSS-Protection header',
                    'References': 'https://developer.mozilla.org/en-US/docs/Web/HTTP/Headers/X-XSS-Protection'
                })
            
            # Check Referrer-Policy
            referrer_policy = response.headers.get('Referrer-Policy')
            if not referrer_policy:
                findings.append({
                    'URL': url,
                    'Type': 'Security Headers',
                    'Name': 'Missing Referrer-Policy',
                    'Description': 'Referrer-Policy header is not set',
                    'Risk': 'Low',
                    'Confidence': 'High',
                    'Path': url,
                    'Solution': 'Implement the Referrer-Policy header',
                    'References': 'https://developer.mozilla.org/en-US/docs/Web/HTTP/Headers/Referrer-Policy'
                })
            
            # Check Feature-Policy / Permissions-Policy
            feature_policy = response.headers.get('Feature-Policy') or response.headers.get('Permissions-Policy')
            if not feature_policy:
                findings.append({
                    'URL': url,
                    'Type': 'Security Headers',
                    'Name': 'Missing Permissions-Policy',
                    'Description': 'Permissions-Policy (formerly Feature-Policy) header is not set',
                    'Risk': 'Low',
                    'Confidence': 'High',
                    'Path': url,
                    'Solution': 'Implement the Permissions-Policy header',
                    'References': 'https://developer.mozilla.org/en-US/docs/Web/HTTP/Headers/Feature-Policy'
                })
            
            # Check for cookies security
            cookies = response.cookies
            for cookie in cookies:
                issues = []
                
                # Check for Secure flag
                if not cookie.secure and parsed_url.scheme == 'https':
                    issues.append('missing Secure flag')
                
                # Check for HttpOnly flag
                if not cookie.has_nonstandard_attr('HttpOnly'):
                    issues.append('missing HttpOnly flag')
                
                # Check for SameSite attribute
                if not cookie.has_nonstandard_attr('SameSite'):
                    issues.append('missing SameSite attribute')
                
                if issues:
                    findings.append({
                        'URL': url,
                        'Type': 'Cookie Security',
                        'Name': f'Insecure Cookie: {cookie.name}',
                        'Description': f"Cookie '{cookie.name}' has {', '.join(issues)}",
                        'Risk': 'Medium',
                        'Confidence': 'High',
                        'Path': url,
                        'Solution': 'Set Secure, HttpOnly, and SameSite attributes on cookies',
                        'References': 'https://owasp.org/www-project-web-security-testing-guide/latest/4-Web_Application_Security_Testing/06-Session_Management_Testing/02-Testing_for_Cookies_Attributes'
                    })
            
            # Check for CORS misconfiguration
            origin = 'https://evil-site.com'
            headers = {'Origin': origin}
            try:
                cors_response = self.session.get(url, headers=headers, timeout=5, verify=not self.args.ignore_ssl)
                acao_header = cors_response.headers.get('Access-Control-Allow-Origin')
                
                if acao_header == '*' or acao_header == origin:
                    findings.append({
                        'URL': url,
                        'Type': 'CORS',
                        'Name': 'CORS Misconfiguration',
                        'Description': f"Server has permissive CORS policy: {acao_header}",
                        'Risk': 'Medium',
                        'Confidence': 'Medium',
                        'Path': url,
                        'Solution': 'Configure CORS to only allow trusted origins',
                        'References': 'https://owasp.org/www-project-web-security-testing-guide/latest/4-Web_Application_Security_Testing/11-Client-side_Testing/07-Testing_Cross_Origin_Resource_Sharing'
                    })
            except:
                pass
            
            # Basic reflected XSS check - just look for reflection, not actual injection
            test_param = 'xsscanary9124'
            test_urls = []
            
            # Build some test URLs with the canary parameter
            parsed = urlparse(url)
            if not parsed.query:
                # If there's no query string, add our test parameter
                test_url = f"{url}?id={test_param}"
                test_urls.append(test_url)
            else:
                # If there's already a query string, try to modify existing parameters
                params = parsed.query.split('&')
                base_url = f"{parsed.scheme}://{parsed.netloc}{parsed.path}"
                
                if len(params) <= 5:  # Only try if there are not too many parameters
                    for param in params:
                        if '=' in param:
                            name, _ = param.split('=', 1)
                            new_params = [p if not p.startswith(f"{name}=") else f"{name}={test_param}" for p in params]
                            test_urls.append(f"{base_url}?{'&'.join(new_params)}")
            
            # Test for reflection
            for test_url in test_urls:
                try:
                    xss_response = self.session.get(test_url, timeout=5, verify=not self.args.ignore_ssl)
                    if test_param in xss_response.text:
                        findings.append({
                            'URL': url,
                            'Type': 'Potential XSS',
                            'Name': 'Reflected Parameter',
                            'Description': f"Parameter value is reflected in the response without encoding",
                            'Risk': 'Medium',
                            'Confidence': 'Low',
                            'Path': test_url,
                            'Solution': 'Implement proper output encoding and input validation',
                            'References': 'https://owasp.org/www-project-web-security-testing-guide/latest/4-Web_Application_Security_Testing/07-Input_Validation_Testing/01-Testing_for_Reflected_Cross_Site_Scripting'
                        })
                        break  # One finding is enough for this check
                except:
                    pass
                
        except Exception as e:
            logger.error(f"Error during basic security checks: {e}")
            
        return findings
        
    def run_scans(self):
        """Run scans on all URLs"""
        self.read_urls()
        
        for url in self.urls:
            try:
                findings = self.scan_url(url)
                self.results.extend(findings)
            except Exception as e:
                logger.error(f"Error scanning {url}: {e}")
                
        logger.info(f"Completed all scans. Found {len(self.results)} issues across {len(self.urls)} URLs")
    
    def generate_report(self):
        """Generate Excel report with findings"""
        if not self.results:
            logger.warning("No results to report")
            return
            
        logger.info(f"Generating report: {self.output_file}")
        
        # Convert results to DataFrame
        df = pd.DataFrame(self.results)
        
        # Create Excel writer
        writer = pd.ExcelWriter(self.output_file, engine='openpyxl')
        
        # Create main findings sheet
        df.to_excel(writer, sheet_name='Security Findings', index=False)
        
        # Get the workbook and the worksheet
        workbook = writer.book
        worksheet = writer.sheets['Security Findings']
        
        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format headers
        header_fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        bold_font = Font(bold=True)
        
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(1, col)
            cell.fill = header_fill
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Adjust column width
            worksheet.column_dimensions[get_column_letter(col)].width = 20
            
            # Set autofilter for the header row
            worksheet.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
        
        # Color rows by risk level with enhanced color coding
        risk_map = {
            'Critical': self.colors['critical'],
            'High': self.colors['high'], 
            'Medium': self.colors['medium'], 
            'Low': self.colors['low'], 
            'Info': self.colors['info']
        }
        
        if 'Risk' in df.columns:
            risk_col = df.columns.get_loc('Risk') + 1  # +1 because openpyxl is 1-indexed
            
            for row in range(2, len(df) + 2):  # +2 because openpyxl is 1-indexed and we have a header row
                risk = worksheet.cell(row, risk_col).value
                fill_color = risk_map.get(risk, self.colors['info'])
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row, col)
                    cell.border = thin_border
                    
                    # Apply different coloring based on severity for entire row
                    if risk in ['Critical', 'High']:
                        if col == risk_col:
                            # Make the risk cell more prominent
                            cell.fill = fill
                            cell.font = Font(bold=True)
                        else:
                            # Lighter shade for other cells in critical/high rows
                            light_fill = PatternFill(
                                start_color='FFF0F0', 
                                end_color='FFF0F0', 
                                fill_type='solid'
                            )
                            cell.fill = light_fill
                    elif col == risk_col:
                        # Only color the risk cell for medium/low/info
                        cell.fill = fill
                    
                    # Wrap text for all cells
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Add table for filtering
        table_ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
        table = Table(displayName="SecurityFindings", ref=table_ref)
        
        # Add a default style to the table
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        
        worksheet.add_table(table)
        
        # Add summary sheet grouped by Risk and Type
        summary_df = df.groupby(['Risk', 'Type']).size().reset_index(name='Count')
        summary_df = summary_df.sort_values(['Risk', 'Count'], ascending=[True, False])
        
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        summary_sheet = writer.sheets['Summary']
        
        # Format summary sheet
        for col in range(1, len(summary_df.columns) + 1):
            cell = summary_sheet.cell(1, col)
            cell.fill = header_fill
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            
            summary_sheet.column_dimensions[get_column_letter(col)].width = 15
            
        # Color summary rows by risk
        if 'Risk' in summary_df.columns:
            risk_col = summary_df.columns.get_loc('Risk') + 1
            
            for row in range(2, len(summary_df) + 2):
                risk = summary_sheet.cell(row, risk_col).value
                fill_color = risk_map.get(risk, self.colors['info'])
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                
                for col in range(1, len(summary_df.columns) + 1):
                    cell = summary_sheet.cell(row, col)
                    cell.border = thin_border
                    
                    if col == risk_col:
                        cell.fill = fill
                        
        # Add URL summary sheet
        url_summary_df = df.groupby(['URL', 'Risk']).size().reset_index(name='Issues')
        url_pivot = url_summary_df.pivot_table(
            index='URL', 
            columns='Risk', 
            values='Issues',
            aggfunc='sum',
            fill_value=0
        ).reset_index()
        
        # Ensure all risk levels are represented
        for risk in ['Critical', 'High', 'Medium', 'Low', 'Info']:
            if risk not in url_pivot.columns:
                url_pivot[risk] = 0
                
        # Calculate total issues per URL
        url_pivot['Total'] = url_pivot.sum(axis=1, numeric_only=True)
        
        # Reorder columns with URL first, then risks from highest to lowest
        url_pivot = url_pivot[['URL', 'Critical', 'High', 'Medium', 'Low', 'Info', 'Total']]
        
        # Sort by total issues descending
        url_pivot = url_pivot.sort_values('Total', ascending=False)
        
        # Write to Excel
        url_pivot.to_excel(writer, sheet_name='URL Summary', index=False)
        url_sheet = writer.sheets['URL Summary']
        
        # Format URL summary sheet
        for col in range(1, len(url_pivot.columns) + 1):
            cell = url_sheet.cell(1, col)
            cell.fill = header_fill
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            
            url_sheet.column_dimensions[get_column_letter(col)].width = 20 if col == 1 else 15
            
        # Color risk columns in the header
        risk_cols = {
            'Critical': 2,
            'High': 3,
            'Medium': 4,
            'Low': 5,
            'Info': 6
        }
        
        for risk, col in risk_cols.items():
            if col <= len(url_pivot.columns):
                cell = url_sheet.cell(1, col)
                fill_color = risk_map.get(risk, self.colors['info'])
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        
        # Save the workbook
        writer.close()
        logger.info(f"Report saved to {self.output_file}")

def parse_arguments():
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(description='DAST Scanner - Dynamic Application Security Testing Tool')
    
    parser.add_argument('-i', '--input-file', required=True, help='Input file containing URLs to scan (one per line)')
    parser.add_argument('-o', '--output', help='Output Excel file for the report')
    parser.add_argument('--zap-proxy', help='OWASP ZAP proxy address (e.g., http://localhost:8080)')
    parser.add_argument('--zap-api-key', help='OWASP ZAP API key')
    parser.add_argument('--require-zap', action='store_true', help='Exit if ZAP connection fails')
    parser.add_argument('--ignore-ssl', action='store_true', help='Ignore SSL certificate errors')
    
    return parser.parse_args()

def main():
    """Main entry point"""
    args = parse_arguments()
    
    print("""
    ██████╗  █████╗ ███████╗████████╗    ███████╗ ██████╗ █████╗ ███╗   ██╗███╗   ██╗███████╗██████╗ 
    ██╔══██╗██╔══██╗██╔════╝╚══██╔══╝    ██╔════╝██╔════╝██╔══██╗████╗  ██║████╗  ██║██╔════╝██╔══██╗
    ██║  ██║███████║███████╗   ██║       ███████╗██║     ███████║██╔██╗ ██║██╔██╗ ██║█████╗  ██████╔╝
    ██║  ██║██╔══██║╚════██║   ██║       ╚════██║██║     ██╔══██║██║╚██╗██║██║╚██╗██║██╔══╝  ██╔══██╗
    ██████╔╝██║  ██║███████║   ██║       ███████║╚██████╗██║  ██║██║ ╚████║██║ ╚████║███████╗██║  ██║
    ╚═════╝ ╚═╝  ╚═╝╚══════╝   ╚═╝       ╚══════╝ ╚═════╝╚═╝  ╚═╝╚═╝  ╚═══╝╚═╝  ╚═══╝╚══════╝╚═╝  ╚═╝                                                         
    Dynamic Application Security Testing Scanner
    """)
    
    scanner = DASTScanner(args)
    scanner.run_scans()
    scanner.generate_report()
    
    print(f"\nScan complete! Report saved to: {scanner.output_file}")

if __name__ == "__main__":
    main() 