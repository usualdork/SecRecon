#!/usr/bin/env python3

import re
import os
import sys
import time
import argparse
import concurrent.futures
import urllib.parse
from datetime import datetime
from urllib.robotparser import RobotFileParser

import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.worksheet.filters import FilterColumn, Filters

# Suppress InsecureRequestWarning
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

class CredentialScanner:
    def __init__(self, domains_file, output_file=None, max_depth=3, threads=10, timeout=10, delay=1):
        self.domains_file = domains_file
        self.output_file = output_file or f"credential_scan_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.max_depth = max_depth
        self.threads = threads
        self.timeout = timeout
        self.delay = delay
        
        # Patterns to search for
        self.patterns = {
            'email': r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}',
            'phone': r'\b(?:\+\d{1,2}\s?)?\(?\d{3}\)?[\s.-]?\d{3}[\s.-]?\d{4}\b',
            'api_key': r'(?i)(api[_-]?key|apikey|access[_-]?key|auth[_-]?token)[\s:=]+[\'"](\w+)[\'"]',
            'aws_key': r'(?i)AKIA[0-9A-Z]{16}',
            'aws_secret': r'(?i)[\'"][0-9a-zA-Z/+]{40}[\'"\s]',
            'google_api': r'(?i)AIza[0-9A-Za-z\-_]{35}',
            'jwt_token': r'eyJ[a-zA-Z0-9_-]*\.[a-zA-Z0-9_-]*\.[a-zA-Z0-9_-]*',
            'password': r'(?i)password[\s:=]+[\'"](\w+)[\'"]',
            'private_key': r'-----BEGIN [A-Z]+ PRIVATE KEY-----',
            'secret_key': r'(?i)(secret[_-]?key|secretkey)[\s:=]+[\'"](\w+)[\'"]',
            'slack_token': r'xox[pbar]-[0-9]{12}-[0-9]{12}-[0-9]{12}-[a-z0-9]{32}',
            'stripe_key': r'(?i)sk_live_[0-9a-zA-Z]{24}',
            'github_token': r'(?i)github[_-]?token[\s:=]+[\'"](\w+)[\'"]',
            'firebase_key': r'(?i)AAAA[A-Za-z0-9_-]{7}:[A-Za-z0-9_-]{140}',
            'ssh_key': r'ssh-rsa AAAA[0-9A-Za-z+/]+[=]{0,3}',
            'username': r'(?i)username[\s:=]+[\'"](\w+)[\'"]',
            'credit_card': r'\b(?:\d{4}[- ]?){3}\d{4}\b',
            'social_security': r'\b\d{3}-\d{2}-\d{4}\b',
        }
        
        # Results storage
        self.results = []
        self.visited_urls = set()
        self.robots_cache = {}
        
        # User agents
        self.user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        
        # Headers
        self.headers = {
            'User-Agent': self.user_agent,
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.5',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
        }
    
    def load_domains(self):
        """Load domains from the file"""
        with open(self.domains_file, 'r') as f:
            return [line.strip() for line in f if line.strip()]
    
    def check_robots_txt(self, domain):
        """Check robots.txt for allowed/disallowed paths"""
        if domain in self.robots_cache:
            return self.robots_cache[domain]
        
        rp = RobotFileParser()
        try:
            robots_url = f"https://{domain}/robots.txt"
            response = requests.get(robots_url, headers=self.headers, timeout=self.timeout, verify=False)
            if response.status_code == 200:
                rp.parse(response.text.splitlines())
                self.robots_cache[domain] = rp
                return rp
        except Exception as e:
            print(f"Error fetching robots.txt for {domain}: {e}")
        
        # If no robots.txt or error, create an empty one that allows everything
        rp = RobotFileParser()
        rp.allow_all = True
        self.robots_cache[domain] = rp
        return rp
    
    def is_allowed(self, url, rp):
        """Check if URL is allowed by robots.txt"""
        if rp.allow_all:
            return True
        return rp.can_fetch(self.user_agent, url)
    
    def normalize_url(self, url, base_url):
        """Normalize URL to absolute form"""
        if not url:
            return None
        
        # Skip non-http links, anchors, javascript, etc.
        if url.startswith(('#', 'javascript:', 'mailto:', 'tel:')):
            return None
        
        # Convert to absolute URL
        absolute_url = urllib.parse.urljoin(base_url, url)
        
        # Remove fragments
        absolute_url = absolute_url.split('#')[0]
        
        # Ensure URL is from the same domain
        base_domain = urllib.parse.urlparse(base_url).netloc
        url_domain = urllib.parse.urlparse(absolute_url).netloc
        
        if base_domain != url_domain:
            return None
        
        return absolute_url
    
    def extract_links(self, soup, base_url):
        """Extract links from HTML"""
        links = []
        for a_tag in soup.find_all('a', href=True):
            url = self.normalize_url(a_tag['href'], base_url)
            if url and url not in self.visited_urls:
                links.append(url)
        return links
    
    def scan_content(self, url, content, domain):
        """Scan content for sensitive information"""
        findings = []
        
        for pattern_name, pattern in self.patterns.items():
            matches = re.findall(pattern, content)
            if matches:
                # If matches is a list of tuples (from capturing groups), flatten it
                if isinstance(matches[0], tuple):
                    matches = [match[1] if len(match) > 1 else match[0] for match in matches]
                
                # Remove duplicates while preserving order
                unique_matches = []
                for match in matches:
                    if match not in unique_matches:
                        unique_matches.append(match)
                
                for match in unique_matches:
                    findings.append({
                        'domain': domain,
                        'url': url,
                        'type': pattern_name,
                        'value': match,
                        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    })
        
        return findings
    
    def crawl_url(self, url, domain, depth=0, rp=None):
        """Crawl a URL and scan for sensitive information"""
        if depth > self.max_depth or url in self.visited_urls:
            return []
        
        self.visited_urls.add(url)
        
        if not rp:
            rp = self.check_robots_txt(domain)
        
        if not self.is_allowed(url, rp):
            print(f"Skipping {url} (disallowed by robots.txt)")
            return []
        
        try:
            print(f"Scanning {url} (depth {depth})")
            response = requests.get(url, headers=self.headers, timeout=self.timeout, verify=False)
            
            # Skip non-text content
            content_type = response.headers.get('Content-Type', '')
            if not content_type.startswith(('text/', 'application/json', 'application/javascript', 'application/xml')):
                return []
            
            content = response.text
            findings = self.scan_content(url, content, domain)
            
            # If HTML, extract links for further crawling
            if content_type.startswith('text/html'):
                soup = BeautifulSoup(content, 'html.parser')
                links = self.extract_links(soup, url)
                
                # Respect rate limiting
                time.sleep(self.delay)
                
                # Crawl extracted links
                if depth < self.max_depth:
                    for link in links[:20]:  # Limit to 20 links per page to prevent explosion
                        findings.extend(self.crawl_url(link, domain, depth + 1, rp))
            
            return findings
        
        except Exception as e:
            print(f"Error crawling {url}: {e}")
            return []
    
    def scan_domain(self, domain):
        """Scan a single domain"""
        print(f"\nScanning domain: {domain}")
        url = f"https://{domain}"
        rp = self.check_robots_txt(domain)
        findings = self.crawl_url(url, domain, 0, rp)
        return findings
    
    def scan_domains(self):
        """Scan all domains"""
        domains = self.load_domains()
        print(f"Loaded {len(domains)} domains for scanning")
        
        all_findings = []
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=self.threads) as executor:
            future_to_domain = {executor.submit(self.scan_domain, domain): domain for domain in domains}
            
            for future in concurrent.futures.as_completed(future_to_domain):
                domain = future_to_domain[future]
                try:
                    findings = future.result()
                    all_findings.extend(findings)
                    print(f"Found {len(findings)} potential issues in {domain}")
                except Exception as e:
                    print(f"Error scanning {domain}: {e}")
        
        self.results = all_findings
        return all_findings
    
    def generate_excel_report(self):
        """Generate a professional Excel report with the findings"""
        if not self.results:
            print("No results to report.")
            return
        
        # Create a pandas DataFrame
        df = pd.DataFrame(self.results)
        
        # Create a new workbook
        workbook = Workbook()
        
        # Remove default sheet
        if 'Sheet' in workbook.sheetnames:
            default_sheet = workbook['Sheet']
            workbook.remove(default_sheet)
        
        # Create summary sheet
        summary_sheet = workbook.create_sheet(title='Summary')
        
        # Add summary information
        summary_sheet['A1'] = 'Credential Scanner Report'
        summary_sheet['A1'].font = Font(size=16, bold=True)
        summary_sheet['A3'] = 'Scan Date:'
        summary_sheet['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        summary_sheet['A4'] = 'Domains Scanned:'
        summary_sheet['B4'] = len(set(df['domain']) if not df.empty else 0)
        summary_sheet['A5'] = 'Total Findings:'
        summary_sheet['B5'] = len(df) if not df.empty else 0
        
        # Add findings by type
        if not df.empty:
            findings_by_type = df['type'].value_counts().reset_index()
            findings_by_type.columns = ['Finding Type', 'Count']
            
            summary_sheet['A7'] = 'Findings by Type'
            summary_sheet['A7'].font = Font(bold=True)
            
            for i, (finding_type, count) in enumerate(zip(findings_by_type['Finding Type'], findings_by_type['Count'])):
                summary_sheet[f'A{8+i}'] = finding_type
                summary_sheet[f'B{8+i}'] = count
        
        # Create findings sheet
        if not df.empty:
            # Group findings by type
            for finding_type in df['type'].unique():
                type_df = df[df['type'] == finding_type].copy()
                sheet_name = finding_type[:31]  # Excel sheet names limited to 31 chars
                
                # Create the sheet
                sheet = workbook.create_sheet(title=sheet_name)
                
                # Write headers
                for col_idx, col_name in enumerate(type_df.columns, 1):
                    cell = sheet.cell(row=1, column=col_idx, value=col_name)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(bold=True, color="FFFFFF")
                
                # Write data
                for row_idx, row in enumerate(type_df.values, 2):
                    for col_idx, value in enumerate(row, 1):
                        cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                        
                        # Add border
                        thin_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        cell.border = thin_border
                        
                        # Highlight sensitive values with different colors based on type
                        if col_idx == type_df.columns.get_loc('type') + 1:
                            cell_value = cell.value
                            if cell_value in ['password', 'api_key', 'secret_key', 'aws_key', 'aws_secret', 'private_key']:
                                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                                cell.font = Font(color="FFFFFF")
                            elif cell_value in ['email', 'username']:
                                cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                            elif cell_value in ['phone', 'credit_card', 'social_security']:
                                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                
                # Format column widths
                for col_idx in range(1, len(type_df.columns) + 1):
                    column_letter = get_column_letter(col_idx)
                    sheet.column_dimensions[column_letter].width = 20
                
                # Add table with filters
                table_name = f"Table_{finding_type}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
                table_name = re.sub(r'[^a-zA-Z0-9_]', '_', table_name)  # Ensure valid table name
                table_ref = f"A1:{get_column_letter(len(type_df.columns))}{len(type_df)+1}"
                
                table = Table(displayName=table_name, ref=table_ref)
                style = TableStyleInfo(
                    name="TableStyleMedium9",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False
                )
                table.tableStyleInfo = style
                sheet.add_table(table)
        
        # Create all findings sheet
        if not df.empty:
            # Create the sheet
            all_sheet = workbook.create_sheet(title='All Findings')
            
            # Write headers
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = all_sheet.cell(row=1, column=col_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold=True, color="FFFFFF")
            
            # Write data
            for row_idx, row in enumerate(df.values, 2):
                for col_idx, value in enumerate(row, 1):
                    cell = all_sheet.cell(row=row_idx, column=col_idx, value=value)
            
            # Format column widths
            for col_idx in range(1, len(df.columns) + 1):
                column_letter = get_column_letter(col_idx)
                all_sheet.column_dimensions[column_letter].width = 20
            
            # Add table with filters
            table_name = f"Table_All_Findings_{datetime.now().strftime('%Y%m%d%H%M%S')}"
            table_name = re.sub(r'[^a-zA-Z0-9_]', '_', table_name)  # Ensure valid table name
            table_ref = f"A1:{get_column_letter(len(df.columns))}{len(df)+1}"
            
            table = Table(displayName=table_name, ref=table_ref)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            all_sheet.add_table(table)
        
        # Save the workbook directly without using pandas ExcelWriter
        workbook.save(self.output_file)
        print(f"\nReport generated: {self.output_file}")

def main():
    parser = argparse.ArgumentParser(description='Credential Scanner - Scan domains for leaked credentials and sensitive information')
    parser.add_argument('domains_file', help='File containing list of domains to scan')
    parser.add_argument('-o', '--output', help='Output Excel file name')
    parser.add_argument('-d', '--depth', type=int, default=3, help='Maximum crawl depth (default: 3)')
    parser.add_argument('-t', '--threads', type=int, default=10, help='Number of concurrent threads (default: 10)')
    parser.add_argument('--timeout', type=int, default=10, help='Request timeout in seconds (default: 10)')
    parser.add_argument('--delay', type=float, default=1, help='Delay between requests in seconds (default: 1)')
    
    args = parser.parse_args()
    
    scanner = CredentialScanner(
        domains_file=args.domains_file,
        output_file=args.output,
        max_depth=args.depth,
        threads=args.threads,
        timeout=args.timeout,
        delay=args.delay
    )
    
    print("Starting credential scan...")
    scanner.scan_domains()
    scanner.generate_excel_report()
    print("Scan completed!")

if __name__ == "__main__":
    main()