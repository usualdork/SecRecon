#!/usr/bin/env python3
# OWASP Top 10 Scanner
# This script scans a list of URLs for OWASP Top 10 vulnerabilities and generates a beautiful Excel report

import argparse
import os
import time
import logging
import sys
import concurrent.futures
import requests
from urllib.parse import urlparse
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import colorama
from colorama import Fore, Style
import threading
import psutil
import signal
import json
import os.path
import traceback
from pathlib import Path
import math

# Initialize colorama for cross-platform colored terminal output
colorama.init()

# Configure logging with colors and better formatting
class ColoredFormatter(logging.Formatter):
    """Custom formatter to add colors to log messages based on level"""
    
    COLORS = {
        'DEBUG': Fore.BLUE,
        'INFO': Fore.GREEN,
        'WARNING': Fore.YELLOW,
        'ERROR': Fore.RED,
        'CRITICAL': Fore.RED + Style.BRIGHT
    }
    
    def format(self, record):
        levelname = record.levelname
        message = super().format(record)
        
        if levelname in self.COLORS:
            message = self.COLORS[levelname] + message + Style.RESET_ALL
            
        return message

# Create log directory if it doesn't exist
logs_dir = Path("logs")
logs_dir.mkdir(exist_ok=True)

# Create a timestamped log file
log_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
log_file = logs_dir / f"owasp_scan_{log_timestamp}.log"
state_file = logs_dir / f"scan_state_{log_timestamp}.json"

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file)
    ]
)
logger = logging.getLogger('owasp_scanner')

# Add colored console handler
console_handler = logging.StreamHandler()
console_handler.setFormatter(ColoredFormatter('%(asctime)s - %(levelname)s - %(message)s'))
logger.addHandler(console_handler)

# Log startup message
logger.info(f"OWASP Scanner started - Log file: {log_file}")
logger.info(f"Python version: {sys.version}")
logger.info(f"Platform: {sys.platform}")

# Global variables for monitoring
scan_stats = {
    'urls_total': 0,
    'urls_completed': 0,
    'urls_failed': 0,
    'vulnerabilities_found': 0,
    'start_time': None,
    'current_url': None,
    'current_scan': None,
    'scan_timeout': 60 * 60  # 1 hour default timeout
}

# Create state file for resumable scanning
STATE_FILE = "owasp_scan_state.json"

def save_state():
    """Save scan state to a file for possible resumption"""
    with open(STATE_FILE, 'w') as f:
        json.dump({
            'scan_stats': scan_stats,
            'completed_urls': completed_urls,
            'pending_urls': pending_urls,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }, f)
    logger.info(f"Scan state saved to {STATE_FILE}")

def load_state():
    """Load scan state from file if available"""
    global scan_stats, completed_urls, pending_urls
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, 'r') as f:
                state = json.load(f)
                scan_stats = state.get('scan_stats', scan_stats)
                completed_urls = state.get('completed_urls', [])
                pending_urls = state.get('pending_urls', [])
            logger.info(f"Loaded previous scan state from {STATE_FILE}")
            return True
        except Exception as e:
            logger.error(f"Failed to load state file: {e}")
    return False

# Monitor system resources
def monitor_resources():
    """Monitor system resources during scanning"""
    global monitoring
    
    if not psutil:
        return
    
    monitoring = True
    while monitoring:
        try:
            cpu_percent = psutil.cpu_percent(interval=1)
            memory = psutil.virtual_memory()
            memory_percent = memory.percent
            
            # Only log if resource usage is high or every 5 minutes
            if cpu_percent > 80 or memory_percent > 80 or int(time.time()) % 300 < 2:
                logger.info(f"Resource usage - CPU: {cpu_percent}% | Memory: {memory_percent}%")
                
                # Check if ZAP process exists and get its resource usage
                for proc in psutil.process_iter(['pid', 'name', 'cpu_percent', 'memory_info']):
                    if 'zap' in proc.info['name'].lower() or 'java' in proc.info['name'].lower():
                        proc_cpu = proc.info['cpu_percent']
                        proc_mem = proc.memory_info().rss / (1024 * 1024)  # MB
                        logger.info(f"ZAP process: {proc.info['name']} (PID: {proc.info['pid']}) - CPU: {proc_cpu}% | Memory: {proc_mem:.1f} MB")
                
        except Exception as e:
            logger.debug(f"Error in resource monitoring: {e}")
            
        time.sleep(10)

# Signal handler for graceful shutdown
def signal_handler(sig, frame):
    logger.warning("Received interrupt signal. Performing graceful shutdown...")
    save_state()
    logger.info("Scanner state saved. You can resume this scan later.")
    sys.exit(0)

# Register signal handler
signal.signal(signal.SIGINT, signal_handler)

# Monitoring thread flag
monitoring = True

# Lists to track URL scanning status
completed_urls = []
pending_urls = []
failed_urls = []

# Try to import ZAP API
try:
    # Try the newer package name first
    from zaproxy import ZAPv2
    logger.info("Using modern zaproxy package")
except ImportError:
    try:
        # Try the older package name
        from zapv2 import ZAPv2
        logger.info("Using legacy zapv2 package")
    except ImportError:
        logger.error("Python OWASP ZAP API not found. Installing zaproxy package...")
        try:
            import subprocess
            subprocess.check_call([sys.executable, "-m", "pip", "install", "zaproxy"])
            from zaproxy import ZAPv2
            logger.info("Successfully installed and imported zaproxy package")
        except Exception as e:
            logger.error(f"Failed to install zaproxy package: {e}")
            logger.error("Please install it manually: pip install zaproxy")
            sys.exit(1)

# OWASP Top 10 (2021) vulnerabilities and their descriptions
OWASP_TOP_10 = {
    'A01:2021-Broken Access Control': 'Security controls that enforce restrictions on what authenticated users are allowed to do.',
    'A02:2021-Cryptographic Failures': 'Failures related to cryptography that often lead to sensitive data exposure or system compromise.',
    'A03:2021-Injection': 'Code injection vulnerabilities like SQL, NoSQL, OS, and LDAP injection.',
    'A04:2021-Insecure Design': 'Flaws in design and architecture that cannot be fixed by proper implementation.',
    'A05:2021-Security Misconfiguration': 'Missing or incorrect configurations that might expose unnecessary features or sensitive data.',
    'A06:2021-Vulnerable and Outdated Components': 'Using components with known vulnerabilities or outdated software.',
    'A07:2021-Identification and Authentication Failures': 'Incorrect implementation of authentication allowing attackers to compromise passwords or session tokens.',
    'A08:2021-Software and Data Integrity Failures': 'Software and data integrity failures related to code and infrastructure lacking protection.',
    'A09:2021-Security Logging and Monitoring Failures': 'Insufficient logging, detection, monitoring, and response capabilities.',
    'A10:2021-Server-Side Request Forgery': 'SSRF flaws occur when a web application fetches a remote resource without validating the user-supplied URL.'
}

class OWASPScanner:
    def __init__(self, zap_path=None, api_key=None, zap_port=8080, start_zap=False, simulation_mode=False):
        """Initialize the OWASP Scanner with ZAP configuration."""
        self.zap_path = zap_path
        self.zap_port = zap_port
        self.api_key = api_key or ''
        self.zap = None
        self.scan_results = []
        self.simulation_mode = simulation_mode
        self.auto_start_zap = start_zap
        self.user_started_zap = False
        
    def start_zap(self):
        """Start ZAP if path is provided, otherwise connect to running instance."""
        if self.simulation_mode:
            logger.info("Running in simulation mode - no actual ZAP connection needed")
            return True
            
        if self.zap_path:
            logger.info(f"Attempting to start ZAP from: {self.zap_path}")
            import subprocess
            try:
                subprocess.Popen([
                    self.zap_path, 
                    '-daemon', 
                    '-port', str(self.zap_port), 
                    '-config', 'api.key=' + self.api_key
                ])
                # Wait for ZAP to start
                logger.info(f"Waiting for ZAP to start on port {self.zap_port}...")
                time.sleep(15)  # Increased wait time
            except Exception as e:
                logger.error(f"Failed to start ZAP: {e}")
                return False
        else:
            logger.info(f"No ZAP path provided, attempting to connect to running ZAP instance on port {self.zap_port}")
            
        # Connect to ZAP
        try:
            proxies = {'http': f'http://localhost:{self.zap_port}', 'https': f'http://localhost:{self.zap_port}'}
            logger.info(f"Connecting to ZAP with proxies: {proxies}")
            
            self.zap = ZAPv2(apikey=self.api_key, proxies=proxies)
            version = self.zap.core.version
            logger.info(f"Successfully connected to ZAP {version}")
            self.user_started_zap = True
            return True
        except Exception as e:
            logger.error(f"Failed to connect to ZAP: {e}")
            logger.error(f"Please make sure ZAP is running on port {self.zap_port} with API key '{self.api_key}'")
            logger.error("To start ZAP manually, run: /Applications/ZAP.app/Contents/Java/zap.sh -daemon -port 8080")
            return False
            
    def normalize_url(self, url):
        """Ensure URL has a scheme (http/https)."""
        if not url.startswith(('http://', 'https://')):
            # First try HTTPS, as it's more secure
            try:
                response = requests.head(f"https://{url}", timeout=5)
                return f"https://{url}"
            except requests.exceptions.RequestException:
                # Fall back to HTTP if HTTPS fails
                return f"http://{url}"
        return url
            
    def scan_url(self, url):
        """Scan a single URL for OWASP Top 10 vulnerabilities."""
        global scan_stats
        
        if not self.zap and not self.simulation_mode:
            logger.error("ZAP not connected. Cannot scan.")
            return None
            
        normalized_url = self.normalize_url(url)
        logger.info(f"{Fore.CYAN}[SCAN START] {normalized_url}{Style.RESET_ALL}")
        
        # Update scan stats
        scan_stats['current_url'] = normalized_url
        
        # Track scan start time for timeout enforcement
        scan_start_time = time.time()
        
        if self.simulation_mode:
            # Generate simulated results
            logger.info(f"Simulation mode: Generating mock results for {normalized_url}")
            return self._generate_simulated_results(normalized_url)
            
        try:
            # Set a watchdog timer to avoid freezing
            def watchdog_timer():
                if scan_stats['current_url'] == normalized_url:
                    logger.critical(f"SCAN TIMEOUT - {normalized_url} scan has been running for over {scan_stats['scan_timeout']/60} minutes")
                    logger.critical("ZAP may be frozen. Consider restarting the scanner.")
            
            # Set watchdog to alert after timeout period
            timer = threading.Timer(scan_stats['scan_timeout'], watchdog_timer)
            timer.daemon = True
            timer.start()
            
            # Access the URL through ZAP
            logger.info(f"Accessing URL through ZAP: {normalized_url}")
            try:
                self.zap.urlopen(normalized_url)
            except Exception as e:
                logger.error(f"Failed to access URL through ZAP: {e}")
                logger.warning("Continuing with scan despite access issue")
            
            # Spider the site to discover content
            logger.info(f"Starting spider scan for {normalized_url}")
            try:
                scan_id = self.zap.spider.scan(normalized_url)
                logger.info(f"Spider scan ID: {scan_id}")
                
                # Wait for spider to complete with progress updates
                last_progress = -1
                while int(self.zap.spider.status(scan_id)) < 100:
                    current_progress = int(self.zap.spider.status(scan_id))
                    
                    # Only log when progress changes
                    if current_progress > last_progress:
                        logger.info(f"Spider progress: {current_progress}%")
                        last_progress = current_progress
                        
                    # Check for timeout
                    if time.time() - scan_start_time > scan_stats['scan_timeout']:
                        logger.warning(f"Spider timeout for {normalized_url}. Moving on.")
                        break
                        
                    time.sleep(5)
                
                logger.info(f"Spider scan completed for {normalized_url}")
                
            except Exception as e:
                logger.error(f"Spider scan failed: {e}")
                logger.warning("Continuing with active scan despite spider failure")
                
            # Perform active scan
            try:
                logger.info(f"Starting active scan for {normalized_url}")
                scan_id = self.zap.ascan.scan(normalized_url)
                logger.info(f"Active scan ID: {scan_id}")
                scan_stats['current_scan'] = scan_id
                
                # Wait for active scan to complete with progress updates
                last_progress = -1
                last_status_time = time.time()
                
                while int(self.zap.ascan.status(scan_id)) < 100:
                    current_progress = int(self.zap.ascan.status(scan_id))
                    current_time = time.time()
                    
                    # Log progress changes
                    if current_progress > last_progress:
                        # Get currently running plugin if available
                        try:
                            scan_info = self.zap.ascan.scan_progress(scan_id)
                            current_plugin = "Unknown"
                            if isinstance(scan_info, dict) and 'scanProgress' in scan_info:
                                progress_data = scan_info['scanProgress']
                                if len(progress_data) > 0 and 'plugin' in progress_data[0]:
                                    current_plugin = progress_data[0]['plugin']
                            
                            logger.info(f"Active scan progress: {current_progress}% | Current plugin: {current_plugin}")
                        except:
                            logger.info(f"Active scan progress: {current_progress}%")
                            
                        last_progress = current_progress
                        last_status_time = current_time
                    
                    # Log periodic status even if no progress change
                    elif current_time - last_status_time > 300:  # 5 minutes with no progress change
                        logger.warning(f"Active scan stalled at {current_progress}% for 5 minutes")
                        last_status_time = current_time
                        
                    # Check for timeout
                    if time.time() - scan_start_time > scan_stats['scan_timeout']:
                        logger.warning(f"Active scan timeout for {normalized_url}. Moving on.")
                        break
                        
                    time.sleep(10)
                
                logger.info(f"Active scan completed for {normalized_url}")
                
            except Exception as e:
                logger.error(f"Active scan error: {e}")
                
            # Get the alerts
            try:
                logger.info(f"Retrieving alerts for {normalized_url}")
                alerts = self.zap.core.alerts(baseurl=normalized_url)
                logger.info(f"Retrieved {len(alerts)} alerts")
                
            except Exception as e:
                logger.error(f"Error retrieving alerts: {e}")
                alerts = []
                
            # Cancel watchdog timer
            timer.cancel()
            
            # Process the alerts
            findings = self._process_alerts(normalized_url, alerts)
            
            # Update scan statistics
            scan_stats['urls_completed'] += 1
            scan_stats['vulnerabilities_found'] += len(findings.get('vulnerabilities', []))
            completed_urls.append(normalized_url)
            
            # Log completion
            alert_count = len(findings.get('vulnerabilities', []))
            if alert_count > 0:
                logger.info(f"{Fore.YELLOW}[SCAN COMPLETE] {normalized_url} - Found {alert_count} vulnerabilities{Style.RESET_ALL}")
            else:
                logger.info(f"{Fore.GREEN}[SCAN COMPLETE] {normalized_url} - No vulnerabilities found{Style.RESET_ALL}")
                
            return findings
            
        except Exception as e:
            logger.error(f"{Fore.RED}Error scanning {normalized_url}: {e}{Style.RESET_ALL}")
            failed_urls.append(normalized_url)
            scan_stats['urls_failed'] += 1
            
            # Create detailed error report
            error_details = {
                'error_message': str(e),
                'error_type': type(e).__name__,
                'traceback': traceback.format_exc()
            }
            
            return {
                'url': normalized_url,
                'status': 'Error',
                'error_details': error_details,
                'vulnerabilities': []
            }
            
    def _generate_simulated_results(self, url):
        """Generate simulated scan results for demonstration."""
        import random
        domain = urlparse(url).netloc
        
        # Create a random number of vulnerabilities (0-5)
        num_vulns = random.randint(0, 5)
        vulnerabilities = []
        
        # List of possible vulnerabilities
        possible_vulns = [
            {
                'owasp_category': 'A03:2021-Injection',
                'risk_level': 'High',
                'confidence': 'Medium',
                'name': 'SQL Injection',
                'description': 'SQL injection may be possible.',
                'solution': 'Parameterize queries or use an ORM.',
                'evidence': 'Error: You have an error in your SQL syntax',
                'url_instance': f'{url}/login'
            },
            {
                'owasp_category': 'A02:2021-Cryptographic Failures',
                'risk_level': 'Medium',
                'confidence': 'High',
                'name': 'Weak TLS Configuration',
                'description': 'The site is using outdated TLS protocols.',
                'solution': 'Update to TLS 1.3.',
                'evidence': 'TLS 1.0 detected',
                'url_instance': url
            },
            {
                'owasp_category': 'A05:2021-Security Misconfiguration',
                'risk_level': 'Medium',
                'confidence': 'High',
                'name': 'Default Configuration',
                'description': 'Default server configuration detected.',
                'solution': 'Remove default configurations.',
                'evidence': 'Default welcome page found',
                'url_instance': url
            },
            {
                'owasp_category': 'A01:2021-Broken Access Control',
                'risk_level': 'High',
                'confidence': 'Medium',
                'name': 'Improper Access Control',
                'description': 'The application does not properly restrict access to resources.',
                'solution': 'Implement proper access controls.',
                'evidence': 'Direct object reference detected',
                'url_instance': f'{url}/profile/123'
            },
            {
                'owasp_category': 'A07:2021-Identification and Authentication Failures',
                'risk_level': 'Medium',
                'confidence': 'Medium',
                'name': 'Weak Password Policy',
                'description': 'The site allows weak passwords.',
                'solution': 'Implement a stronger password policy.',
                'evidence': 'Password "password" accepted',
                'url_instance': f'{url}/register'
            },
            {
                'owasp_category': 'A09:2021-Security Logging and Monitoring Failures',
                'risk_level': 'Low',
                'confidence': 'High',
                'name': 'Insufficient Logging',
                'description': 'The application does not log security-related events.',
                'solution': 'Implement proper logging.',
                'evidence': 'No logging detected for failed login attempts',
                'url_instance': f'{url}/login'
            }
        ]
        
        # Add random vulnerabilities
        for _ in range(num_vulns):
            vuln = random.choice(possible_vulns)
            vulnerabilities.append(vuln.copy())  # Create a copy to avoid modifying the original
            
        return {
            'url': url,
            'domain': domain,
            'status': 'Completed',
            'vulnerabilities': vulnerabilities,
            'scan_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
    
    def _map_to_owasp_top10(self, alert):
        """Map ZAP alert to OWASP Top 10 category."""
        # This is a simplified mapping - a real implementation would be more comprehensive
        alert_name = alert.get('name', '').lower()
        alert_desc = alert.get('description', '').lower()
        
        # Simplified mapping logic - in a real scenario, this would be much more detailed
        if any(term in alert_name or term in alert_desc for term in ['access control', 'authorization', 'permission']):
            return 'A01:2021-Broken Access Control'
        elif any(term in alert_name or term in alert_desc for term in ['crypto', 'tls', 'ssl', 'encrypt', 'cipher']):
            return 'A02:2021-Cryptographic Failures'
        elif any(term in alert_name or term in alert_desc for term in ['inject', 'sql', 'xss', 'script']):
            return 'A03:2021-Injection'
        elif any(term in alert_name or term in alert_desc for term in ['design', 'architecture']):
            return 'A04:2021-Insecure Design'
        elif any(term in alert_name or term in alert_desc for term in ['config', 'default', 'error message']):
            return 'A05:2021-Security Misconfiguration'
        elif any(term in alert_name or term in alert_desc for term in ['component', 'outdated', 'version', 'library']):
            return 'A06:2021-Vulnerable and Outdated Components'
        elif any(term in alert_name or term in alert_desc for term in ['auth', 'password', 'session', 'login', 'logout']):
            return 'A07:2021-Identification and Authentication Failures'
        elif any(term in alert_name or term in alert_desc for term in ['integrity', 'deserial', 'json', 'xml']):
            return 'A08:2021-Software and Data Integrity Failures'
        elif any(term in alert_name or term in alert_desc for term in ['log', 'monitor', 'audit']):
            return 'A09:2021-Security Logging and Monitoring Failures'
        elif any(term in alert_name or term in alert_desc for term in ['ssrf', 'request forgery', 'server-side']):
            return 'A10:2021-Server-Side Request Forgery'
        
        # Default if no specific mapping is found
        return 'Other'
            
    def _process_alerts(self, url, alerts):
        """Process ZAP alerts into structured data."""
        domain = urlparse(url).netloc
        vulnerabilities = []
        
        for alert in alerts:
            risk_level = alert.get('risk')
            confidence = alert.get('confidence', 'Unknown')
            name = alert.get('name', 'Unknown')
            description = alert.get('description', '')
            solution = alert.get('solution', '')
            evidence = alert.get('evidence', '')
            url_instance = alert.get('url', '')
            
            # Map to OWASP Top 10
            owasp_category = self._map_to_owasp_top10(alert)
            
            vulnerabilities.append({
                'owasp_category': owasp_category,
                'risk_level': risk_level,
                'confidence': confidence,
                'name': name,
                'description': description,
                'solution': solution,
                'evidence': evidence,
                'url_instance': url_instance
            })
            
        return {
            'url': url,
            'domain': domain,
            'status': 'Completed',
            'vulnerabilities': vulnerabilities,
            'scan_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
            
    def scan_urls(self, urls, max_workers=3):
        """Scan multiple URLs with better concurrency control and progress tracking."""
        global scan_stats, pending_urls, monitoring
        
        # Initialize scan statistics
        scan_stats['urls_total'] = len(urls)
        scan_stats['start_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        pending_urls = list(urls)
        results = []
        
        # Print scan header
        logger.info(f"{Fore.CYAN}{'='*80}{Style.RESET_ALL}")
        logger.info(f"{Fore.CYAN}OWASP TOP 10 SCAN STARTED - {len(urls)} URLs{Style.RESET_ALL}")
        logger.info(f"{Fore.CYAN}{'='*80}{Style.RESET_ALL}")
        
        # Check for resumable scan
        if load_state():
            # Filter out already completed URLs
            urls = [url for url in urls if url not in completed_urls]
            logger.info(f"Resuming scan with {len(urls)} remaining URLs")
            
        # Start resource monitoring thread
        monitor_thread = threading.Thread(target=monitor_resources)
        monitor_thread.daemon = True
        monitor_thread.start()
        
        try:
            with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
                # Submit all URLs for scanning
                future_to_url = {executor.submit(self.scan_url, url): url for url in urls}
                
                # Process results as they complete
                for i, future in enumerate(concurrent.futures.as_completed(future_to_url)):
                    url = future_to_url[future]
                    try:
                        result = future.result()
                        if result:
                            results.append(result)
                            try:
                                pending_urls.remove(url)
                            except ValueError:
                                pass  # URL might have been removed already
                            
                            # Print progress
                            completed = i + 1
                            logger.info(f"Progress: {completed}/{len(urls)} URLs ({completed/len(urls)*100:.1f}%)")
                            
                            # Save state periodically
                            if completed % 5 == 0 or completed == len(urls):
                                save_state()
                                
                        else:
                            logger.warning(f"No results for {url}")
                            failed_urls.append(url)
                            scan_stats['urls_failed'] += 1
                            
                    except Exception as e:
                        logger.error(f"Error processing {url}: {e}")
                        logger.error(traceback.format_exc())
                        failed_urls.append(url)
                        scan_stats['urls_failed'] += 1
                
        except KeyboardInterrupt:
            logger.warning("Scan interrupted by user")
            save_state()
            
        except Exception as e:
            logger.error(f"Scan process error: {e}")
            logger.error(traceback.format_exc())
            save_state()
            
        finally:
            # Stop resource monitoring
            monitoring = False
            
            # Print scan summary
            logger.info(f"{Fore.CYAN}{'='*80}{Style.RESET_ALL}")
            logger.info(f"{Fore.CYAN}SCAN SUMMARY{Style.RESET_ALL}")
            logger.info(f"{Fore.CYAN}{'='*80}{Style.RESET_ALL}")
            logger.info(f"Total URLs: {scan_stats['urls_total']}")
            logger.info(f"Completed: {scan_stats['urls_completed']}")
            logger.info(f"Failed: {scan_stats['urls_failed']}")
            logger.info(f"Vulnerabilities found: {scan_stats['vulnerabilities_found']}")
            
            # List failed URLs if any
            if failed_urls:
                logger.warning("Failed URLs:")
                for url in failed_urls:
                    logger.warning(f" - {url}")
            
            self.scan_results = results
            
        return results
    
    def shutdown(self):
        """Shutdown ZAP if it was started by this script."""
        if self.zap and self.zap_path:
            try:
                self.zap.core.shutdown()
                logger.info("ZAP has been shut down")
            except Exception as e:
                logger.error(f"Error shutting down ZAP: {e}")

    def clear_scan_session(self):
        """Clear the current ZAP session to prepare for a new batch of URLs."""
        if self.simulation_mode:
            logger.info("Simulation mode: Pretending to clear ZAP session")
            return True
            
        try:
            # First make sure we're connected
            if not self.zap:
                logger.warning("ZAP not connected, nothing to clear")
                return False
                
            logger.info("Clearing ZAP session")
            # Optionally clear the spider and scan results
            self.zap.spider.remove_all_scans()
            self.zap.ascan.remove_all_scans()
            self.zap.core.new_session()
            return True
        except Exception as e:
            logger.error(f"Error clearing ZAP session: {e}")
            return False

class ExcelReportGenerator:
    """Enhanced Excel report generator with improved formatting and more details"""
    
    def __init__(self, results):
        self.results = results
        self.wb = Workbook()
        self.ws_summary = self.wb.active
        self.ws_summary.title = "Summary"
        self.ws_vulnerabilities = self.wb.create_sheet("Vulnerabilities")
        
        # Define colors and styles
        self.header_fill = PatternFill(start_color="0072BA", end_color="0072BA", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF")
        self.severity_colors = {
            "High": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
            "Medium": PatternFill(start_color="FFAA00", end_color="FFAA00", fill_type="solid"),
            "Low": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
            "Informational": PatternFill(start_color="00AAFF", end_color="00AAFF", fill_type="solid")
        }
        self.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
    def generate_report(self, output_path):
        """Generate a comprehensive Excel report with summary and detailed findings"""
        try:
            # Add report metadata
            self._add_metadata()
            
            # Create summary sheet
            self._create_summary_sheet()
            
            # Create vulnerabilities sheet
            self._create_vulnerabilities_sheet()
            
            # Auto-adjust column widths
            self._adjust_column_widths()
            
            # Save the workbook
            logger.info(f"Saving report to {output_path}")
            self.wb.save(output_path)
            return output_path
        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")
            logger.error(traceback.format_exc())
            return False
            
    def _add_metadata(self):
        """Add metadata and styling to the report"""
        # Add report title
        self.ws_summary.merge_cells('A1:F1')
        title_cell = self.ws_summary['A1']
        title_cell.value = "OWASP Top 10 Vulnerability Scan Report"
        title_cell.font = Font(size=16, bold=True, color="0072BA")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add scan metadata
        self.ws_summary['A3'] = "Scan Date:"
        self.ws_summary['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.ws_summary['A4'] = "URLs Scanned:"
        self.ws_summary['B4'] = len(self.results)
        self.ws_summary['A5'] = "Total Vulnerabilities:"
        
        # Count total vulnerabilities
        total_vulns = sum(len(result.get('vulnerabilities', [])) for result in self.results)
        self.ws_summary['B5'] = total_vulns
        
        # Style the metadata section
        for row in range(3, 6):
            self.ws_summary[f'A{row}'].font = Font(bold=True)
            
    def _create_summary_sheet(self):
        """Create the summary sheet with vulnerability statistics"""
        # Add header row for summary table
        headers = ["URL", "High", "Medium", "Low", "Informational", "Total"]
        self.ws_summary.append([])  # Empty row for spacing
        self.ws_summary.append(headers)
        
        # Apply header styling
        header_row = 7  # Header is at row 7
        for col, header in enumerate(headers, start=1):
            cell = self.ws_summary.cell(row=header_row, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data rows
        row_num = header_row + 1
        total_counts = {"High": 0, "Medium": 0, "Low": 0, "Informational": 0, "Total": 0}
        
        for result in self.results:
            url = result.get('url', 'Unknown')
            vulnerabilities = result.get('vulnerabilities', [])
            
            # Count by severity
            severity_counts = {"High": 0, "Medium": 0, "Low": 0, "Informational": 0}
            for vuln in vulnerabilities:
                severity = vuln.get('risk_level', 'Informational')
                severity_counts[severity] = severity_counts.get(severity, 0) + 1
            
            total = sum(severity_counts.values())
            
            # Update total counts
            for severity, count in severity_counts.items():
                total_counts[severity] += count
            total_counts["Total"] += total
            
            # Add row to sheet
            row_data = [url]
            for severity in ["High", "Medium", "Low", "Informational"]:
                row_data.append(severity_counts.get(severity, 0))
            row_data.append(total)
            
            self.ws_summary.append(row_data)
            
            # Apply cell styling
            for col in range(1, len(headers) + 1):
                cell = self.ws_summary.cell(row=row_num, column=col)
                cell.border = self.border
                if col > 1 and col < len(headers):  # Severity columns
                    severity = headers[col-1]
                    if severity_counts.get(severity, 0) > 0:
                        cell.fill = self.severity_colors.get(severity, None)
                
            row_num += 1
        
        # Add totals row
        self.ws_summary.append([])
        totals_row = ["TOTAL"]
        for severity in ["High", "Medium", "Low", "Informational", "Total"]:
            totals_row.append(total_counts.get(severity, 0))
        
        self.ws_summary.append(totals_row)
        
        # Style totals row
        for col in range(1, len(headers) + 1):
            cell = self.ws_summary.cell(row=row_num + 1, column=col)
            cell.font = Font(bold=True)
            cell.border = self.border
            if col > 1 and col < len(headers):  # Severity columns
                severity = headers[col-1]
                if total_counts.get(severity, 0) > 0:
                    cell.fill = self.severity_colors.get(severity, None)
        
    def _create_vulnerabilities_sheet(self):
        """Create the detailed vulnerabilities sheet"""
        # Add headers
        headers = [
            "URL", "Vulnerability", "OWASP Category", "Severity", "CWE ID", 
            "Description", "Solution", "Parameter", "Evidence"
        ]
        
        self.ws_vulnerabilities.append(headers)
        
        # Style headers
        for col, header in enumerate(headers, start=1):
            cell = self.ws_vulnerabilities.cell(row=1, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add vulnerability data
        row_num = 2
        for result in self.results:
            url = result.get('url', 'Unknown')
            vulnerabilities = result.get('vulnerabilities', [])
            
            for vuln in vulnerabilities:
                name = vuln.get('name', 'Unknown')
                category = vuln.get('owasp_category', 'Unknown')
                severity = vuln.get('risk_level', 'Informational')
                cwe_id = vuln.get('cwe_id', 'N/A')
                description = vuln.get('description', 'No description available')
                solution = vuln.get('solution', 'No solution provided')
                parameter = vuln.get('param', 'N/A')
                evidence = vuln.get('evidence', 'N/A')
                
                # Add row
                self.ws_vulnerabilities.append([
                    url, name, category, severity, cwe_id, 
                    description, solution, parameter, evidence
                ])
                
                # Apply styling
                for col in range(1, len(headers) + 1):
                    cell = self.ws_vulnerabilities.cell(row=row_num, column=col)
                    cell.border = self.border
                    
                    # Apply severity color
                    if col == 4:  # Severity column
                        cell.fill = self.severity_colors.get(severity, None)
                
                # Multi-line text alignment
                for col in [6, 7]:  # Description and Solution columns
                    cell = self.ws_vulnerabilities.cell(row=row_num, column=col)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                row_num += 1
    
    def _adjust_column_widths(self):
        """Auto-adjust column widths based on content"""
        for ws in [self.ws_summary, self.ws_vulnerabilities]:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = min(adjusted_width, 50)

def read_urls_from_file(file_path):
    """Read URLs from a file, one URL per line."""
    with open(file_path, 'r') as f:
        return [line.strip() for line in f if line.strip()]

def split_into_batches(urls, batch_size):
    """Split a list of URLs into smaller batches."""
    return [urls[i:i + batch_size] for i in range(0, len(urls), batch_size)]

def main():
    """Main function to handle command line arguments and execute the scanner."""
    parser = argparse.ArgumentParser(description='OWASP Top 10 Vulnerability Scanner using ZAP')
    
    # Input options
    parser.add_argument('--url', help='Single URL to scan')
    parser.add_argument('--url-file', help='File containing URLs to scan (one per line)')
    
    # ZAP configuration
    parser.add_argument('--zap-path', help='Path to ZAP executable (e.g., /path/to/zap.sh)')
    parser.add_argument('--zap-port', type=int, default=8080, help='Port for ZAP API (default: 8080)')
    parser.add_argument('--api-key', default='api-key-for-owasp', help='ZAP API key')
    parser.add_argument('--start-zap', action='store_true', help='Start ZAP automatically if not running')
    parser.add_argument('--simulation-mode', action='store_true', help='Run in simulation mode (no actual scanning)')
    
    # Scanner options
    parser.add_argument('--max-workers', type=int, default=3, help='Maximum number of concurrent scans (default: 3)')
    parser.add_argument('--scan-timeout', type=int, default=60, help='Timeout in minutes for each URL scan (default: 60)')
    parser.add_argument('--resume', action='store_true', help='Resume from last saved state if available')
    parser.add_argument('--batch-size', type=int, default=3, help='Number of URLs to scan in each batch (default: 3)')
    
    # Output options
    parser.add_argument('--output', default='owasp_scan_results.xlsx', help='Output Excel file path')
    parser.add_argument('--log-level', choices=['DEBUG', 'INFO', 'WARNING', 'ERROR'], default='INFO', 
                        help='Set the logging level (default: INFO)')
    
    args = parser.parse_args()
    
    # Set log level
    log_level = getattr(logging, args.log_level)
    logger.setLevel(log_level)
    
    # Set global scan timeout
    scan_stats['scan_timeout'] = args.scan_timeout * 60  # Convert minutes to seconds
    
    # Validate inputs
    if not args.url and not args.url_file:
        logger.error("Either --url or --url-file must be specified")
        parser.print_help()
        sys.exit(1)
        
    # Get URLs to scan
    urls = []
    if args.url:
        urls.append(args.url)
    elif args.url_file:
        try:
            urls = read_urls_from_file(args.url_file)
        except Exception as e:
            logger.error(f"Error reading URL file: {e}")
            sys.exit(1)
            
    if not urls:
        logger.error("No valid URLs found to scan")
        sys.exit(1)
        
    logger.info(f"Preparing to scan {len(urls)} URLs")
    
    # Create and configure the scanner
    try:
        scanner = OWASPScanner(
            zap_path=args.zap_path,
            api_key=args.api_key,
            zap_port=args.zap_port,
            start_zap=args.start_zap,
            simulation_mode=args.simulation_mode
        )
        
        # Start ZAP if it's not in simulation mode
        if not args.simulation_mode:
            if not scanner.start_zap():
                logger.error("Failed to start or connect to ZAP. Exiting.")
                sys.exit(1)
        
        # Process URLs in batches if there are more than batch_size URLs
        all_results = []
        
        if len(urls) > args.batch_size:
            logger.info(f"Processing {len(urls)} URLs in batches of {args.batch_size}")
            batches = split_into_batches(urls, args.batch_size)
            total_batches = len(batches)
            
            for i, batch in enumerate(batches):
                logger.info(f"Processing batch {i+1}/{total_batches} with {len(batch)} URLs")
                batch_results = scanner.scan_urls(batch, max_workers=min(args.max_workers, len(batch)))
                
                if batch_results:
                    all_results.extend(batch_results)
                    # Print batch summary
                    logger.info(f"Batch {i+1}/{total_batches} completed:")
                    logger.info(f"  URLs processed: {len(batch)}")
                    vulnerabilities = sum(len(result.get('vulnerabilities', [])) for result in batch_results)
                    logger.info(f"  Vulnerabilities found: {vulnerabilities}")
                else:
                    logger.warning(f"No results from batch {i+1}/{total_batches}")
                
                # Reset the scanner for the next batch if needed
                if i < total_batches - 1:  # If not the last batch
                    logger.info("Clearing scanner state for next batch")
                    scanner.clear_scan_session()
                    time.sleep(2)  # Give ZAP a moment to reset
                
        else:
            # For small URL lists, process normally
            all_results = scanner.scan_urls(urls, max_workers=args.max_workers)
            
        # Handle results
        if not all_results:
            logger.warning("No scan results generated")
            sys.exit(1)
            
        # Generate report
        logger.info(f"Generating Excel report to {args.output}")
        try:
            report_generator = ExcelReportGenerator(all_results)
            report_path = report_generator.generate_report(args.output)
            
            logger.info(f"Report successfully generated: {report_path}")
            
            # Print summary
            scan_stats['end_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            logger.info("Scan completed successfully!")
            logger.info(f"Total URLs: {len(urls)}")
            logger.info(f"Completed: {len(urls) - len(failed_urls)}")
            logger.info(f"Failed: {len(failed_urls)}")
            
            total_vulnerabilities = sum(len(result.get('vulnerabilities', [])) for result in all_results)
            logger.info(f"Vulnerabilities found: {total_vulnerabilities}")
            logger.info(f"Results saved to {args.output}")
            
        except Exception as e:
            logger.error(f"Error generating report: {e}")
            logger.debug(traceback.format_exc())
            sys.exit(1)
            
        # Cleanup
        scanner.shutdown()
        
    except KeyboardInterrupt:
        logger.info("Scan interrupted by user")
        logger.info("Saving current state...")
        save_state()
        sys.exit(0)
    except Exception as e:
        logger.error(f"Error during scan: {e}")
        logger.debug(traceback.format_exc())
        sys.exit(1)

if __name__ == "__main__":
    main() 