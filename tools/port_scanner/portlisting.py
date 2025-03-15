import subprocess
import sys
import os
import re
import time
import xml.etree.ElementTree as ET
from datetime import datetime
import requests
from requests.packages.urllib3.exceptions import InsecureRequestWarning
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import argparse
import concurrent.futures

# Disable SSL warnings
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

class DirectNmapScanner:
    """
    A direct Nmap scanner implementation with minimal processing
    and straightforward web technology detection.
    """
    
    def __init__(self, nmap_path=None, max_workers=3, timeout=300):
        """Initialize the scanner with configuration parameters."""
        self.max_workers = max_workers
        self.timeout = timeout
        self.nmap_path = self._find_nmap_executable(nmap_path)
        
        if not self.nmap_path:
            print("[-] Error: Nmap executable not found. Please install Nmap.")
            sys.exit(1)
            
        print(f"[+] Using Nmap at: {self.nmap_path}")
    
    def _find_nmap_executable(self, specified_path=None):
        """Find the Nmap executable path."""
        if specified_path and os.path.isfile(specified_path) and os.access(specified_path, os.X_OK):
            return specified_path
        
        # Check common paths
        potential_paths = [
            "nmap",  # In PATH
            "/usr/bin/nmap",
            "/usr/local/bin/nmap",
            "C:\\Program Files\\Nmap\\nmap.exe",
            "C:\\Program Files (x86)\\Nmap\\nmap.exe"
        ]
        
        for path in potential_paths:
            try:
                subprocess.run([path, "-V"], stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=2)
                return path
            except (subprocess.SubprocessError, FileNotFoundError, PermissionError):
                continue
        
        return None
    
    def scan_target(self, target):
        """
        Scan a single target using Nmap.
        
        Parameters:
            target (str): The hostname to scan
            
        Returns:
            dict: Scan results
        """
        print(f"\n[+] Scanning {target}")
        result = {
            'hostname': target,
            'ip_address': None,
            'scan_time': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'open_ports': [],
            'os_info': None,
            'technologies': [],
            'error': None
        }
        
        # Use a unique output file for this scan
        timestamp = int(time.time())
        xml_output = f"nmap_scan_{target.replace('.', '_')}_{timestamp}.xml"
        
        try:
            # First run a quick scan to see if host is up and get initial info
            print(f"[+] Running initial Nmap scan on {target}")
            cmd = [
                self.nmap_path,
                "-sS",           # SYN scan
                "-T4",           # Faster timing
                "--open",        # Only show open ports
                "-Pn",           # Treat all hosts as online
                "-n",            # No DNS resolution
                "-oX", xml_output,
                "--top-ports", "1000",  # Scan most common ports first
                target
            ]
            
            init_process = subprocess.run(
                cmd, 
                stdout=subprocess.PIPE, 
                stderr=subprocess.PIPE, 
                text=True,
                timeout=self.timeout
            )
            
            # Parse initial scan results
            if os.path.exists(xml_output) and os.path.getsize(xml_output) > 0:
                initial_results = self._parse_nmap_xml(xml_output)
                
                # If host seems up and has open ports, do a more comprehensive scan
                if initial_results and initial_results.get('open_ports'):
                    print(f"[+] Host {target} appears active with {len(initial_results['open_ports'])} open ports in initial scan")
                    print(f"[+] Running comprehensive scan on {target}")
                    
                    # Try to do a more comprehensive scan with version detection
                    cmd = [
                        self.nmap_path,
                        "-sS",           # SYN scan
                        "-sV",           # Version detection
                        "-A",            # OS detection, version detection, script scanning, and traceroute
                        "-T4",           # Faster timing
                        "--open",        # Only show open ports
                        "-Pn",           # Treat all hosts as online
                        "-n",            # No DNS resolution
                        "-oX", xml_output,
                        "--version-intensity", "7",  # More aggressive version detection
                        "--max-retries", "2",
                        target
                    ]
                    
                    full_process = subprocess.run(
                        cmd, 
                        stdout=subprocess.PIPE, 
                        stderr=subprocess.PIPE, 
                        text=True,
                        timeout=self.timeout
                    )
                    
                    # Parse final results
                    if os.path.exists(xml_output) and os.path.getsize(xml_output) > 0:
                        final_results = self._parse_nmap_xml(xml_output)
                        if final_results:
                            result.update(final_results)
                            
                            # Try to detect web technologies on HTTP/HTTPS ports
                            web_techs = self._detect_web_technologies(target, result['open_ports'])
                            result['technologies'] = web_techs
                else:
                    if initial_results:
                        result.update(initial_results)
                        print(f"[+] Host {target} appears to have no open ports in initial scan")
                    else:
                        result['error'] = "Failed to parse initial scan results"
                        print(f"[-] Failed to parse initial scan results for {target}")
            else:
                result['error'] = "No XML output generated from Nmap"
                print(f"[-] No XML output generated from Nmap for {target}")
                
            # Clean up the XML file
            try:
                if os.path.exists(xml_output):
                    os.remove(xml_output)
            except:
                pass
                
        except subprocess.TimeoutExpired:
            result['error'] = f"Scan timed out after {self.timeout} seconds"
            print(f"[-] Scan timed out for {target}")
        except Exception as e:
            result['error'] = f"Error: {str(e)}"
            print(f"[-] Error scanning {target}: {str(e)}")
        
        return result
        
    def _parse_nmap_xml(self, xml_file):
        """
        Parse Nmap XML output.
        
        Parameters:
            xml_file (str): Path to XML file
            
        Returns:
            dict: Parsed results
        """
        try:
            tree = ET.parse(xml_file)
            root = tree.getroot()
            
            results = {
                'ip_address': None,
                'open_ports': [],
                'os_info': None
            }
            
            # Find host element
            host_elem = root.find(".//host")
            if host_elem is None:
                return None
                
            # Get IP address
            addr_elem = host_elem.find(".//address[@addrtype='ipv4']")
            if addr_elem is not None:
                results['ip_address'] = addr_elem.get('addr')
                
            # Get open ports
            ports_elem = host_elem.find(".//ports")
            if ports_elem is not None:
                for port_elem in ports_elem.findall(".//port"):
                    state_elem = port_elem.find(".//state")
                    if state_elem is not None and state_elem.get('state') == 'open':
                        port_info = {
                            'port': int(port_elem.get('portid')),
                            'protocol': port_elem.get('protocol', 'tcp'),
                            'service': None,
                            'product': None,
                            'version': None,
                            'extra_info': None
                        }
                        
                        # Get service info
                        service_elem = port_elem.find(".//service")
                        if service_elem is not None:
                            port_info['service'] = service_elem.get('name', 'unknown')
                            port_info['product'] = service_elem.get('product')
                            port_info['version'] = service_elem.get('version')
                            port_info['extra_info'] = service_elem.get('extrainfo')
                            
                        results['open_ports'].append(port_info)
            
            # Get OS detection
            os_elem = host_elem.find(".//os/osmatch")
            if os_elem is not None:
                results['os_info'] = {
                    'name': os_elem.get('name'),
                    'accuracy': os_elem.get('accuracy')
                }
                
            return results
            
        except ET.ParseError:
            print(f"[-] XML parse error for {xml_file}")
            return None
        except Exception as e:
            print(f"[-] Error parsing Nmap results: {str(e)}")
            return None
            
    def _detect_web_technologies(self, hostname, open_ports):
        """
        Detect web technologies by analyzing HTTP responses.
        
        Parameters:
            hostname (str): Target hostname
            open_ports (list): List of port dictionaries
            
        Returns:
            list: Detected technologies
        """
        technologies = []
        
        # Find HTTP/HTTPS ports
        web_ports = []
        for port_info in open_ports:
            port = port_info['port']
            service = port_info.get('service', '').lower()
            
            if port == 80 or port == 443 or 'http' in service:
                web_ports.append((port, 'https' if port == 443 else 'http'))
                
        # If no web ports found, check common ports anyway
        if not web_ports:
            web_ports = [(80, 'http'), (443, 'https')]
            
        # Simple headers for HTTP requests
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/90.0.4430.212 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.5',
            'Connection': 'close'
        }
        
        # Check each port
        for port, protocol in web_ports:
            try:
                url = f"{protocol}://{hostname}"
                if (protocol == 'http' and port != 80) or (protocol == 'https' and port != 443):
                    url = f"{url}:{port}"
                    
                print(f"[+] Checking web technologies at {url}")
                
                response = requests.get(
                    url,
                    headers=headers,
                    timeout=10,
                    verify=False,
                    allow_redirects=True
                )
                
                if response.status_code < 400:  # Success or redirect
                    # Check for technologies in headers
                    server = response.headers.get('Server')
                    if server:
                        technologies.append({
                            'name': server,
                            'type': 'Server',
                            'url': url
                        })
                        
                    powered_by = response.headers.get('X-Powered-By')
                    if powered_by:
                        technologies.append({
                            'name': powered_by,
                            'type': 'Framework',
                            'url': url
                        })
                        
                    # Check common technology signatures in content
                    content = response.text.lower()
                    
                    # Web frameworks and libraries
                    tech_signatures = {
                        'react': {
                            'patterns': ['reactroot', 'reactdom', 'react.', '_react', '__REACT_'],
                            'type': 'JavaScript Framework'
                        },
                        'angular': {
                            'patterns': ['ng-app', 'ng-controller', 'angular.js', 'angularjs'],
                            'type': 'JavaScript Framework'
                        },
                        'vue.js': {
                            'patterns': ['vue.js', 'vuejs', 'v-app', 'v-bind', 'vue@'],
                            'type': 'JavaScript Framework'
                        },
                        'jquery': {
                            'patterns': ['jquery', 'jquery.min.js'],
                            'type': 'JavaScript Library'
                        },
                        'bootstrap': {
                            'patterns': ['bootstrap.', 'bootstrap.min.', 'bootstrap.css'],
                            'type': 'CSS Framework'
                        },
                        'tailwind': {
                            'patterns': ['tailwind.', 'tailwindcss'],
                            'type': 'CSS Framework'
                        },
                        'wordpress': {
                            'patterns': ['/wp-content/', 'wp-includes', 'wordpress'],
                            'type': 'CMS'
                        },
                        'php': {
                            'patterns': ['.php', 'php'],
                            'type': 'Server Language'
                        },
                        'laravel': {
                            'patterns': ['laravel', 'laravel.', '/laravel/'],
                            'type': 'PHP Framework'
                        },
                        'django': {
                            'patterns': ['django', 'csrftoken', 'csrfmiddlewaretoken'],
                            'type': 'Python Framework'
                        },
                        'flask': {
                            'patterns': ['flask.', 'flask/'],
                            'type': 'Python Framework'
                        },
                        'node.js': {
                            'patterns': ['node', 'nodejs', 'express'],
                            'type': 'Server Runtime'
                        },
                        'next.js': {
                            'patterns': ['__next', 'next/', '_next/'],
                            'type': 'React Framework'
                        },
                        'nuxt.js': {
                            'patterns': ['__nuxt', 'nuxt/', '_nuxt/'],
                            'type': 'Vue Framework'
                        }
                    }
                    
                    for tech, data in tech_signatures.items():
                        for pattern in data['patterns']:
                            if pattern in content:
                                # Check if we already detected this technology
                                if not any(t.get('name', '').lower() == tech for t in technologies):
                                    technologies.append({
                                        'name': tech.capitalize(),
                                        'type': data['type'],
                                        'url': url
                                    })
                                break
                                
                    # Look for version information
                    version_patterns = {
                        'jquery': r'jquery[/-]([0-9.]+)(?:\.min)?\.js',
                        'bootstrap': r'bootstrap[/-]([0-9.]+)(?:\.min)?\.(?:js|css)',
                        'react': r'react[/-]([0-9.]+)(?:\.min)?\.js',
                        'angular': r'angular[/-]([0-9.]+)(?:\.min)?\.js',
                        'vue': r'vue[/-]([0-9.]+)(?:\.min)?\.js'
                    }
                    
                    for tech, pattern in version_patterns.items():
                        match = re.search(pattern, content)
                        if match:
                            version = match.group(1)
                            # Update existing entry or add new one
                            for t in technologies:
                                if t.get('name', '').lower() == tech:
                                    t['version'] = version
                                    break
                            else:
                                technologies.append({
                                    'name': tech.capitalize(),
                                    'type': 'JavaScript Library/Framework',
                                    'version': version,
                                    'url': url
                                })
                
            except requests.RequestException:
                continue
                
        return technologies
        
    def scan_multiple_targets(self, targets):
        """
        Scan multiple targets in parallel.
        
        Parameters:
            targets (list): List of target hostnames
            
        Returns:
            list: Scan results for all targets
        """
        results = []
        start_time = time.time()
        
        print(f"[+] Scanning {len(targets)} targets with {self.max_workers} concurrent workers")
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            future_to_target = {executor.submit(self.scan_target, target): target for target in targets}
            
            for i, future in enumerate(concurrent.futures.as_completed(future_to_target)):
                target = future_to_target[future]
                try:
                    result = future.result()
                    results.append(result)
                    
                    # Report completion status
                    elapsed = time.time() - start_time
                    open_ports_count = len(result.get('open_ports', []))
                    tech_count = len(result.get('technologies', []))
                    
                    print(f"[+] Completed {i+1}/{len(targets)} - {target} - "
                          f"{open_ports_count} open ports - {tech_count} technologies - {elapsed:.1f}s elapsed")
                    
                except Exception as e:
                    print(f"[-] Error scanning {target}: {str(e)}")
                    
        total_elapsed = time.time() - start_time
        print(f"[+] All scans completed in {total_elapsed:.1f} seconds")
        
        return results
        
    def export_to_excel(self, results, output_file):
        """
        Export results to an Excel spreadsheet.
        
        Parameters:
            results (list): Scan results
            output_file (str): Output file path
        """
        print(f"[+] Exporting results to {output_file}")
        
        wb = openpyxl.Workbook()
        
        # Create sheets
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        ws_ports = wb.create_sheet("Open Ports")
        ws_tech = wb.create_sheet("Technologies")
        
        # Style for headers
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')
        
        # Summary sheet
        summary_headers = [
            'Hostname', 'IP Address', 'Open Ports', 'Web Ports', 
            'Technologies', 'OS Detection', 'Scan Status'
        ]
        
        # Set column widths
        column_widths = [30, 15, 60, 30, 40, 30, 20]
        for i, width in enumerate(column_widths, 1):
            ws_summary.column_dimensions[get_column_letter(i)].width = width
            
        # Add headers
        for col_num, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
        # Add data
        for row_num, result in enumerate(results, 2):
            # Get web ports (HTTP/HTTPS)
            web_ports = []
            for port_info in result.get('open_ports', []):
                port = port_info['port']
                service = port_info.get('service', '').lower()
                
                if port == 80 or port == 443 or 'http' in service:
                    product = f" ({port_info['product']})" if port_info.get('product') else ""
                    version = f" {port_info['version']}" if port_info.get('version') else ""
                    web_ports.append(f"{port}/{service}{product}{version}")
            
            # Get all open ports
            open_ports = []
            for port_info in result.get('open_ports', []):
                port = port_info['port']
                service = port_info.get('service', '').lower()
                product = f" ({port_info['product']})" if port_info.get('product') else ""
                version = f" {port_info['version']}" if port_info.get('version') else ""
                open_ports.append(f"{port}/{service}{product}{version}")
            
            # Get technologies
            techs = []
            for tech in result.get('technologies', []):
                version = f" {tech['version']}" if tech.get('version') else ""
                techs.append(f"{tech['name']}{version}")
            
            # Get OS info
            os_info = "Unknown"
            if result.get('os_info'):
                os_name = result['os_info'].get('name', 'Unknown')
                accuracy = result['os_info'].get('accuracy', 'Unknown')
                os_info = f"{os_name} (Accuracy: {accuracy}%)"
                
            # Fill data
            col = 1
            ws_summary.cell(row=row_num, column=col).value = result['hostname']; col += 1
            ws_summary.cell(row=row_num, column=col).value = result.get('ip_address', 'Unknown'); col += 1
            ws_summary.cell(row=row_num, column=col).value = ", ".join(open_ports) if open_ports else 'None'; col += 1
            ws_summary.cell(row=row_num, column=col).value = ", ".join(web_ports) if web_ports else 'None'; col += 1
            ws_summary.cell(row=row_num, column=col).value = ", ".join(techs) if techs else 'None'; col += 1
            ws_summary.cell(row=row_num, column=col).value = os_info; col += 1
            ws_summary.cell(row=row_num, column=col).value = "Error: " + result['error'] if result.get('error') else 'Success'
            
        # Port details sheet
        port_headers = [
            'Hostname', 'IP Address', 'Port', 'Protocol', 'Service', 
            'Product', 'Version', 'Extra Info'
        ]
        
        # Set column widths
        column_widths = [30, 15, 10, 10, 20, 25, 20, 30]
        for i, width in enumerate(column_widths, 1):
            ws_ports.column_dimensions[get_column_letter(i)].width = width
            
        # Add headers
        for col_num, header in enumerate(port_headers, 1):
            cell = ws_ports.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
        # Add data
        row_num = 2
        for result in results:
            if not result.get('open_ports'):
                # Add a row indicating no open ports
                col = 1
                ws_ports.cell(row=row_num, column=col).value = result['hostname']; col += 1
                ws_ports.cell(row=row_num, column=col).value = result.get('ip_address', 'Unknown'); col += 1
                ws_ports.cell(row=row_num, column=col).value = 'No open ports detected'
                row_num += 1
                continue
                
            for port_info in result['open_ports']:
                col = 1
                ws_ports.cell(row=row_num, column=col).value = result['hostname']; col += 1
                ws_ports.cell(row=row_num, column=col).value = result.get('ip_address', 'Unknown'); col += 1
                ws_ports.cell(row=row_num, column=col).value = port_info['port']; col += 1
                ws_ports.cell(row=row_num, column=col).value = port_info['protocol']; col += 1
                ws_ports.cell(row=row_num, column=col).value = port_info.get('service', 'unknown'); col += 1
                ws_ports.cell(row=row_num, column=col).value = port_info.get('product', ''); col += 1
                ws_ports.cell(row=row_num, column=col).value = port_info.get('version', ''); col += 1
                ws_ports.cell(row=row_num, column=col).value = port_info.get('extra_info', '')
                
                row_num += 1
                
        # Technology details sheet
        tech_headers = [
            'Hostname', 'URL', 'Technology', 'Type', 'Version'
        ]
        
        # Set column widths
        column_widths = [30, 40, 25, 25, 15]
        for i, width in enumerate(column_widths, 1):
            ws_tech.column_dimensions[get_column_letter(i)].width = width
            
        # Add headers
        for col_num, header in enumerate(tech_headers, 1):
            cell = ws_tech.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
        # Add data
        row_num = 2
        for result in results:
            if not result.get('technologies'):
                # Add a row indicating no technologies detected
                col = 1
                ws_tech.cell(row=row_num, column=col).value = result['hostname']; col += 1
                ws_tech.cell(row=row_num, column=col).value = f"https://{result['hostname']}"; col += 1
                ws_tech.cell(row=row_num, column=col).value = 'No technologies detected'
                row_num += 1
                continue
                
            for tech in result['technologies']:
                col = 1
                ws_tech.cell(row=row_num, column=col).value = result['hostname']; col += 1
                ws_tech.cell(row=row_num, column=col).value = tech.get('url', f"https://{result['hostname']}"); col += 1
                ws_tech.cell(row=row_num, column=col).value = tech['name']; col += 1
                ws_tech.cell(row=row_num, column=col).value = tech.get('type', 'Unknown'); col += 1
                ws_tech.cell(row=row_num, column=col).value = tech.get('version', '')
                
                row_num += 1
                
        # Apply filters
        ws_summary.auto_filter.ref = f"A1:{get_column_letter(len(summary_headers))}{len(results) + 1}"
        ws_summary.freeze_panes = 'A2'
        
        ws_ports.auto_filter.ref = f"A1:{get_column_letter(len(port_headers))}{ws_ports.max_row}"
        ws_ports.freeze_panes = 'A2'
        
        ws_tech.auto_filter.ref = f"A1:{get_column_letter(len(tech_headers))}{ws_tech.max_row}"
        ws_tech.freeze_panes = 'A2'
        
        # Save the workbook
        try:
            wb.save(output_file)
            print(f"[+] Results exported to {output_file}")
        except Exception as e:
            print(f"[-] Error saving Excel file: {str(e)}")
            alternative_file = f"scan_results_{int(time.time())}.xlsx"
            print(f"[+] Trying alternative filename: {alternative_file}")
            wb.save(alternative_file)
            print(f"[+] Results exported to {alternative_file}")


def main():
    """Main function to handle command line arguments."""
    parser = argparse.ArgumentParser(
        description="Direct Nmap scanner with web technology detection"
    )
    
    # Input parameters
    input_group = parser.add_mutually_exclusive_group(required=True)
    input_group.add_argument('-t', '--targets', help='Comma-separated list of targets')
    input_group.add_argument('-f', '--file', help='File containing target hostnames (one per line)')
    
    # Output parameters
    parser.add_argument('-o', '--output', help='Output Excel file (default: scan_results_TIMESTAMP.xlsx)')
    
    # Scan parameters
    parser.add_argument('-w', '--workers', type=int, default=3, help='Number of concurrent workers')
    parser.add_argument('-T', '--timeout', type=int, default=300, help='Scan timeout in seconds')
    parser.add_argument('-p', '--nmap-path', help='Path to nmap executable')
    
    args = parser.parse_args()
    
    # Process targets
    targets = []
    if args.targets:
        targets = [t.strip() for t in args.targets.split(',') if t.strip()]
    elif args.file:
        try:
            with open(args.file, 'r') as f:
                targets = [line.strip() for line in f if line.strip()]
        except Exception as e:
            print(f"[-] Error reading target file: {str(e)}")
            sys.exit(1)
    
    if not targets:
        print("[-] No targets specified")
        sys.exit(1)
        
    # Remove duplicates while preserving order
    unique_targets = []
    seen = set()
    for target in targets:
        if target not in seen:
            unique_targets.append(target)
            seen.add(target)
            
    print(f"[+] Loaded {len(unique_targets)} unique targets")
    
    # Generate output filename if not specified
    output_file = args.output
    if not output_file:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"scan_results_{timestamp}.xlsx"
        
    # Initialize scanner and run scans
    scanner = DirectNmapScanner(
        nmap_path=args.nmap_path,
        max_workers=args.workers,
        timeout=args.timeout
    )
    
    # Scan targets
    results = scanner.scan_multiple_targets(unique_targets)
    
    # Export results
    scanner.export_to_excel(results, output_file)


if __name__ == "__main__":
    print("\n" + "=" * 70)
    print("  DIRECT NMAP SCANNER WITH WEB TECHNOLOGY DETECTION")
    print("=" * 70)
    
    try:
        main()
    except KeyboardInterrupt:
        print("\n[-] Scan interrupted by user")
    except Exception as e:
        print(f"\n[-] Unhandled error: {str(e)}")